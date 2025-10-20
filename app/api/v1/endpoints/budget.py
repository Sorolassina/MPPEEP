# app/api/v1/endpoints/budget.py
"""
Endpoints pour la gestion budgétaire et les conférences budgétaires
"""

import io
import re
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO
from pathlib import Path

import pandas as pd
from fastapi import APIRouter, Depends, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse, StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from sqlmodel import Session, delete, func, select

from app.api.v1.endpoints.auth import get_current_user
from app.core.logging_config import get_logger
from app.db.session import get_session
from app.models.budget import (
    ActionBudgetaire,
    Activite,
    ActiviteBudgetaire,
    DocumentBudget,
    DocumentLigneBudgetaire,
    FicheTechnique,
    HistoriqueBudget,
    LigneBudgetaire,
    LigneBudgetaireDetail,
    NatureDepense,
    ServiceBeneficiaire,
    SigobeChargement,
    SigobeExecution,
    SigobeKpi,
)
from app.models.personnel import Direction, Programme
from app.models.user import User
from app.services.activity_service import ActivityService
from app.services.fiche_technique_service import FicheTechniqueService
from app.services.sigobe_service import SigobeService
from app.templates import get_template_context, templates

logger = get_logger(__name__)
router = APIRouter()


# ============================================
# DASHBOARD BUDGÉTAIRE
# ============================================


@router.get("/", response_class=HTMLResponse, name="budget_home")
def budget_home(
    request: Request,
    annee: int | None = None,
    programme_id: int | None = None,
    trimestre: int | None = None,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """
    Dashboard principal du suivi budgétaire
    Utilise les données SIGOBE pour les KPIs
    """
    # Déterminer l'année à utiliser
    if not annee:
        # Si pas d'année spécifiée, chercher la dernière année disponible dans SIGOBE
        dernier_chargement_any = session.exec(
            select(SigobeChargement).order_by(SigobeChargement.annee.desc(), SigobeChargement.date_chargement.desc())
        ).first()

        if dernier_chargement_any:
            annee = dernier_chargement_any.annee
            logger.info(f"📅 Aucune année spécifiée, utilisation de la dernière année SIGOBE : {annee}")
        else:
            annee = datetime.now().year
            logger.warning(f"⚠️ Aucune donnée SIGOBE trouvée, utilisation de l'année courante : {annee}")

    # Récupérer le dernier chargement SIGOBE pour cette année
    chargement_query = select(SigobeChargement).where(SigobeChargement.annee == annee)
    if trimestre:
        chargement_query = chargement_query.where(SigobeChargement.trimestre == trimestre)

    dernier_chargement = session.exec(chargement_query.order_by(SigobeChargement.date_chargement.desc())).first()

    logger.info(
        f"🔍 Dashboard SIGOBE - Année: {annee}, Chargement: {dernier_chargement.id if dernier_chargement else 'Aucun'}"
    )

    # KPIs par défaut
    budget_vote_total = 0
    budget_actuel_total = 0
    engagements_total = 0
    mandats_emis_total = 0
    mandats_vises_total = 0
    mandats_pec_total = 0
    disponible_eng_total = 0
    taux_engagement = 0
    taux_mandatement_emis = 0
    taux_mandatement_vise = 0
    taux_mandatement_pec = 0
    taux_execution_global = 0

    exec_par_programme = {}
    exec_par_nature = {}

    if dernier_chargement:
        # Récupérer les KPIs globaux
        kpi_global = session.exec(
            select(SigobeKpi)
            .where(SigobeKpi.chargement_id == dernier_chargement.id)
            .where(SigobeKpi.dimension == "global")
        ).first()

        if kpi_global:
            budget_vote_total = float(kpi_global.budget_vote_total or 0)
            budget_actuel_total = float(kpi_global.budget_actuel_total or 0)
            engagements_total = float(kpi_global.engagements_total or 0)
            mandats_emis_total = float(kpi_global.mandats_total or 0)

            # Récupérer les totaux détaillés depuis SigobeExecution
            executions_sigobe = session.exec(
                select(SigobeExecution).where(SigobeExecution.chargement_id == dernier_chargement.id)
            ).all()

            mandats_vises_total = sum(float(e.mandats_vise_cf or 0) for e in executions_sigobe)
            mandats_pec_total = sum(float(e.mandats_pec or 0) for e in executions_sigobe)
            disponible_eng_total = sum(float(e.disponible_eng or 0) for e in executions_sigobe)

            # Calculer les taux selon vos formules DAX
            # _Tx_Eng : DIVIDE(Engagements, Budget_Actuel)
            budg_select = budget_actuel_total or budget_vote_total
            taux_engagement = (engagements_total / budg_select * 100) if budg_select > 0 else 0

            # _Tx_Mandat_Emis : DIVIDE(Mandats_Emis, Engagements)
            taux_mandatement_emis = (mandats_emis_total / engagements_total * 100) if engagements_total > 0 else 0

            # _Tx_Mandat_Vise : DIVIDE(Mandats_Vise, Mandats_PEC)
            taux_mandatement_vise = (mandats_vises_total / mandats_pec_total * 100) if mandats_pec_total > 0 else 0

            # _Tx_Mandat_PEC : DIVIDE(Mandats_PEC, Mandats_Emis)
            taux_mandatement_pec = (mandats_pec_total / mandats_emis_total * 100) if mandats_emis_total > 0 else 0

            # _Tx_Exe.Global : DIVIDE(Disponible, Budget_Actuel)
            taux_execution_global = (disponible_eng_total / budg_select * 100) if budg_select > 0 else 0

        # Récupérer les KPIs par programme
        if dernier_chargement:
            kpis_programmes = session.exec(
                select(SigobeKpi)
                .where(SigobeKpi.chargement_id == dernier_chargement.id)
                .where(SigobeKpi.dimension == "programme")
            ).all()

            for kpi in kpis_programmes:
                # Utiliser le libellé comme clé si le code est vide
                code = kpi.dimension_code or kpi.dimension_libelle or "INCONNU"
                exec_par_programme[code] = {
                    "libelle": kpi.dimension_libelle or code,
                    "budget": float(kpi.budget_actuel_total or 0),
                    "engagements": float(kpi.engagements_total or 0),
                    "mandats": float(kpi.mandats_total or 0),
                    "taux": float(kpi.taux_execution or 0),
                }

            # Récupérer les KPIs par nature
            kpis_natures = session.exec(
                select(SigobeKpi)
                .where(SigobeKpi.chargement_id == dernier_chargement.id)
                .where(SigobeKpi.dimension == "nature")
            ).all()

            for kpi in kpis_natures:
                # Utiliser le libellé comme clé si le code est vide
                code = kpi.dimension_code or kpi.dimension_libelle or "INCONNU"
                exec_par_nature[code] = {
                    "libelle": kpi.dimension_libelle or code,
                    "budget": float(kpi.budget_actuel_total or 0),
                    "engagements": float(kpi.engagements_total or 0),
                    "mandats": float(kpi.mandats_total or 0),
                    "taux": float(kpi.taux_execution or 0),
                }

    # Récupérer les programmes pour les filtres
    programmes = session.exec(select(Programme).where(Programme.actif)).all()

    # Récupérer les années disponibles dans SIGOBE
    annees_sigobe_query = select(SigobeChargement.annee).distinct().order_by(SigobeChargement.annee.desc())
    annees_sigobe = [a for a in session.exec(annees_sigobe_query).all()]

    # Calculer les variations par rapport à l'année précédente
    variation_engagement = None
    variation_mandatement_vise = None
    variation_mandatement_pec = None
    variation_execution_global = None
    taux_engagement_n1 = None
    taux_mandatement_vise_n1 = None
    taux_mandatement_pec_n1 = None
    taux_execution_global_n1 = None
    budget_actuel_n1 = 0
    budget_vote_n1 = 0
    engagements_n1 = 0

    if annee and (annee - 1) in annees_sigobe:
        # Récupérer le dernier chargement de l'année N-1
        chargement_n1_query = select(SigobeChargement).where(SigobeChargement.annee == (annee - 1))
        if trimestre:
            chargement_n1_query = chargement_n1_query.where(SigobeChargement.trimestre == trimestre)

        chargement_n1 = session.exec(chargement_n1_query.order_by(SigobeChargement.date_chargement.desc())).first()

        if chargement_n1:
            # Récupérer les KPIs globaux de N-1
            kpi_global_n1 = session.exec(
                select(SigobeKpi)
                .where(SigobeKpi.chargement_id == chargement_n1.id)
                .where(SigobeKpi.dimension == "global")
            ).first()

            if kpi_global_n1:
                # Récupérer les données brutes de N-1 pour appliquer les MÊMES formules que N
                executions_sigobe_n1 = session.exec(
                    select(SigobeExecution).where(SigobeExecution.chargement_id == chargement_n1.id)
                ).all()

                budget_actuel_n1 = float(kpi_global_n1.budget_actuel_total or 0)
                budget_vote_n1 = float(kpi_global_n1.budget_vote_total or 0)
                engagements_n1 = float(kpi_global_n1.engagements_total or 0)
                mandats_emis_n1 = float(kpi_global_n1.mandats_total or 0)

                mandats_vises_n1 = sum(float(e.mandats_vise_cf or 0) for e in executions_sigobe_n1)
                mandats_pec_n1 = sum(float(e.mandats_pec or 0) for e in executions_sigobe_n1)
                disponible_eng_n1 = sum(float(e.disponible_eng or 0) for e in executions_sigobe_n1)

                # Calculer les taux N-1 avec les MÊMES formules que N
                budg_select_n1 = budget_actuel_n1 or budget_vote_n1
                taux_engagement_n1 = (engagements_n1 / budg_select_n1 * 100) if budg_select_n1 > 0 else 0
                taux_mandatement_vise_n1 = (mandats_vises_n1 / mandats_pec_n1 * 100) if mandats_pec_n1 > 0 else 0
                taux_mandatement_pec_n1 = (mandats_pec_n1 / mandats_emis_n1 * 100) if mandats_emis_n1 > 0 else 0
                taux_execution_global_n1 = (disponible_eng_n1 / budg_select_n1 * 100) if budg_select_n1 > 0 else 0

                # Calculer les variations (différence absolue en points de pourcentage)
                variation_engagement = taux_engagement - taux_engagement_n1
                variation_mandatement_vise = taux_mandatement_vise - taux_mandatement_vise_n1
                variation_mandatement_pec = taux_mandatement_pec - taux_mandatement_pec_n1
                variation_execution_global = taux_execution_global - taux_execution_global_n1

    return templates.TemplateResponse(
        "pages/budget_dashboard.html",
        get_template_context(
            request,
            annee=annee,
            trimestre=trimestre,
            annees_disponibles=annees_sigobe,
            budget_vote_total=budget_vote_total,
            budget_actuel_total=budget_actuel_total,
            engagements_total=engagements_total,
            mandats_emis_total=mandats_emis_total,
            mandats_vises_total=mandats_vises_total,
            mandats_pec_total=mandats_pec_total,
            disponible_eng_total=disponible_eng_total,
            taux_engagement=round(taux_engagement, 2),
            taux_mandatement_emis=round(taux_mandatement_emis, 2),
            taux_mandatement_vise=round(taux_mandatement_vise, 2),
            taux_mandatement_pec=round(taux_mandatement_pec, 2),
            taux_execution_global=round(taux_execution_global, 2),
            exec_par_programme=exec_par_programme,
            exec_par_nature=exec_par_nature,
            programmes=programmes,
            dernier_chargement=dernier_chargement,
            # Variations N vs N-1
            variation_engagement=round(variation_engagement, 2) if variation_engagement is not None else None,
            variation_mandatement_vise=round(variation_mandatement_vise, 2)
            if variation_mandatement_vise is not None
            else None,
            variation_mandatement_pec=round(variation_mandatement_pec, 2)
            if variation_mandatement_pec is not None
            else None,
            variation_execution_global=round(variation_execution_global, 2)
            if variation_execution_global is not None
            else None,
            taux_engagement_n1=round(taux_engagement_n1, 2) if taux_engagement_n1 is not None else None,
            taux_mandatement_vise_n1=round(taux_mandatement_vise_n1, 2)
            if taux_mandatement_vise_n1 is not None
            else None,
            taux_mandatement_pec_n1=round(taux_mandatement_pec_n1, 2) if taux_mandatement_pec_n1 is not None else None,
            taux_execution_global_n1=round(taux_execution_global_n1, 2)
            if taux_execution_global_n1 is not None
            else None,
            budget_actuel_n1=budget_actuel_n1,
            budget_vote_n1=budget_vote_n1,
            engagements_n1=engagements_n1,
            current_user=current_user,
        ),
    )


# ============================================
# PAGE D'AIDE
# ============================================


@router.get("/aide", response_class=HTMLResponse, name="aide_budget")
def aide_budget(request: Request):
    """Page d'aide pour le module Budget"""
    return templates.TemplateResponse("pages/aide_budget.html", get_template_context(request))


# ============================================
# FICHES TECHNIQUES
# ============================================


@router.get("/fiches", response_class=HTMLResponse, name="budget_fiches")
def budget_fiches(
    request: Request,
    annee: int | None = None,
):
    """Redirection vers la version hiérarchique des fiches"""
    # Rediriger vers la nouvelle version hiérarchique
    if annee:
        return RedirectResponse(
            url=str(request.url_for("budget_fiches_hierarchique")) + f"?annee={annee}", status_code=303
        )
    else:
        return RedirectResponse(
            url=str(request.url_for("budget_fiches_hierarchique")) + "?annee=toutes", status_code=303
        )


@router.get("/fiches/hierarchique", response_class=HTMLResponse, name="budget_fiches_hierarchique")
def budget_fiches_hierarchique(
    request: Request,
    annee: str | None = None,
    programme_id: str | None = None,
    direction_id: str | None = None,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Liste des fiches techniques avec structure hiérarchique"""
    # Construire la requête avec filtres
    query = select(FicheTechnique).where(FicheTechnique.actif)

    # Filtre par année
    if annee and annee != "toutes":
        query = query.where(FicheTechnique.annee_budget == int(annee))

    # Filtre par programme
    if programme_id and programme_id != "tous":
        query = query.where(FicheTechnique.programme_id == int(programme_id))

    # Filtre par direction
    if direction_id and direction_id != "toutes":
        query = query.where(FicheTechnique.direction_id == int(direction_id))

    fiches = session.exec(query.order_by(FicheTechnique.created_at.desc())).all()

    # Valeurs sélectionnées pour l'affichage
    annee_selectionnee = annee if annee else "toutes"
    programme_selectionne = programme_id if programme_id else "tous"
    direction_selectionnee = direction_id if direction_id else "toutes"

    # Référentiels
    programmes_dict = {p.id: p for p in session.exec(select(Programme)).all()}
    directions_dict = {d.id: d for d in session.exec(select(Direction)).all()}

    # Listes pour les filtres
    programmes_list = session.exec(select(Programme).where(Programme.actif).order_by(Programme.code)).all()
    directions_list = session.exec(select(Direction).where(Direction.actif).order_by(Direction.code)).all()

    # Années disponibles
    annees_disponibles = session.exec(
        select(FicheTechnique.annee_budget).distinct().order_by(FicheTechnique.annee_budget.desc())
    ).all()

    return templates.TemplateResponse(
        "pages/budget_fiches_hierarchique.html",
        get_template_context(
            request,
            annee=annee_selectionnee,
            programme_id=programme_selectionne,
            direction_id=direction_selectionnee,
            fiches=fiches,
            programmes=programmes_dict,
            directions=directions_dict,
            programmes_list=programmes_list,
            directions_list=directions_list,
            annees_disponibles=annees_disponibles,
            current_user=current_user,
        ),
    )


@router.get("/fiches/nouveau", response_class=HTMLResponse, name="budget_fiche_nouveau")
def budget_fiche_nouveau(
    request: Request, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """Formulaire de création de fiche technique"""
    programmes = session.exec(select(Programme).where(Programme.actif).order_by(Programme.code)).all()
    directions = session.exec(select(Direction).where(Direction.actif).order_by(Direction.libelle)).all()
    natures = session.exec(select(NatureDepense).where(NatureDepense.actif).order_by(NatureDepense.code)).all()
    activites = session.exec(select(Activite).where(Activite.actif).order_by(Activite.code)).all()

    return templates.TemplateResponse(
        "pages/budget_fiche_form.html",
        get_template_context(
            request,
            mode="create",
            programmes=programmes,
            directions=directions,
            natures=natures,
            activites=activites,
            current_user=current_user,
        ),
    )


@router.get("/fiches/{fiche_id}/edit", response_class=HTMLResponse, name="budget_fiche_edit")
def budget_fiche_edit(
    request: Request,
    fiche_id: int,
):
    """Redirection vers la structure hiérarchique pour modifier les lignes"""
    # Une fiche technique ne se modifie pas directement, on modifie ses lignes via la structure
    return RedirectResponse(url=str(request.url_for("budget_fiche_structure", fiche_id=fiche_id)), status_code=303)


@router.get("/fiches/{fiche_id}", response_class=HTMLResponse, name="budget_fiche_detail")
def budget_fiche_detail(
    fiche_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Détail d'une fiche technique"""
    fiche = session.get(FicheTechnique, fiche_id)
    if not fiche:
        raise HTTPException(404, "Fiche technique non trouvée")

    # Lignes budgétaires
    lignes = session.exec(
        select(LigneBudgetaire)
        .where(LigneBudgetaire.fiche_technique_id == fiche_id)
        .where(LigneBudgetaire.actif)
        .order_by(LigneBudgetaire.ordre)
    ).all()

    # Documents
    documents = session.exec(
        select(DocumentBudget)
        .where(DocumentBudget.fiche_technique_id == fiche_id)
        .where(DocumentBudget.actif)
        .order_by(DocumentBudget.uploaded_at.desc())
    ).all()

    # Historique
    historique = session.exec(
        select(HistoriqueBudget)
        .where(HistoriqueBudget.fiche_technique_id == fiche_id)
        .order_by(HistoriqueBudget.date_action.desc())
    ).all()

    # Référentiels
    programme = session.get(Programme, fiche.programme_id)
    direction = session.get(Direction, fiche.direction_id) if fiche.direction_id else None
    natures = {n.id: n for n in session.exec(select(NatureDepense)).all()}
    activites = {a.id: a for a in session.exec(select(Activite)).all()}

    return templates.TemplateResponse(
        "pages/budget_fiche_detail.html",
        get_template_context(
            request,
            fiche=fiche,
            lignes=lignes,
            documents=documents,
            historique=historique,
            programme=programme,
            direction=direction,
            natures=natures,
            activites=activites,
            current_user=current_user,
        ),
    )


# ============================================
# API - CRUD FICHES TECHNIQUES
# ============================================


@router.post("/api/fiches", name="api_creer_fiche_vide")
async def api_create_fiche(
    annee_budget: int = Form(...),
    programme_id: int = Form(...),
    direction_id: int | None = Form(None),
    budget_anterieur: float = Form(0),
    enveloppe_demandee: float = Form(...),
    complement_demande: float = Form(0),
    engagement_etat: float = Form(0),
    financement_bailleurs: float = Form(0),
    note_justificative: str | None = Form(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Créer une nouvelle fiche technique"""
    # Générer numéro de fiche
    count = session.exec(select(func.count(FicheTechnique.id)).where(FicheTechnique.annee_budget == annee_budget)).one()

    prog = session.get(Programme, programme_id)
    numero_fiche = f"FT-{annee_budget}-{prog.code}-{count + 1:03d}"

    # Calculer le total
    budget_total = (
        Decimal(str(enveloppe_demandee))
        + Decimal(str(complement_demande))
        + Decimal(str(engagement_etat))
        + Decimal(str(financement_bailleurs))
    )

    fiche = FicheTechnique(
        numero_fiche=numero_fiche,
        annee_budget=annee_budget,
        programme_id=programme_id,
        direction_id=direction_id,
        budget_anterieur=Decimal(str(budget_anterieur)),
        enveloppe_demandee=Decimal(str(enveloppe_demandee)),
        complement_demande=Decimal(str(complement_demande)),
        engagement_etat=Decimal(str(engagement_etat)),
        financement_bailleurs=Decimal(str(financement_bailleurs)),
        budget_total_demande=budget_total,
        note_justificative=note_justificative,
        created_by_user_id=current_user.id,
    )

    session.add(fiche)
    session.commit()
    session.refresh(fiche)

    # Historique
    hist = HistoriqueBudget(
        fiche_technique_id=fiche.id,
        action="Création",
        nouveau_statut="Brouillon",
        montant_apres=budget_total,
        commentaire="Fiche technique créée",
        user_id=current_user.id,
    )
    session.add(hist)
    session.commit()

    logger.info(f"✅ Fiche technique créée : {numero_fiche} par {current_user.email}")

    # Log activité
    ActivityService.log_user_activity(
        session=session,
        user=current_user,
        action_type="create",
        target_type="fiche_technique",
        description=f"Création de la fiche technique {numero_fiche} - Budget {annee_budget}",
        target_id=fiche.id,
        icon="📋",
    )

    return {"ok": True, "id": fiche.id, "numero": numero_fiche}


@router.delete("/api/fiches/{fiche_id}")
def api_delete_fiche(
    fiche_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """Supprimer une fiche technique et toute sa structure hiérarchique"""
    fiche = session.get(FicheTechnique, fiche_id)
    if not fiche:
        raise HTTPException(404, "Fiche non trouvée")

    try:
        # Supprimer dans l'ordre hiérarchique (du bas vers le haut)
        # 1. Supprimer les lignes budgétaires
        session.exec(delete(LigneBudgetaireDetail).where(LigneBudgetaireDetail.fiche_technique_id == fiche_id))

        # 2. Supprimer les activités
        session.exec(delete(ActiviteBudgetaire).where(ActiviteBudgetaire.fiche_technique_id == fiche_id))

        # 3. Supprimer les services
        session.exec(delete(ServiceBeneficiaire).where(ServiceBeneficiaire.fiche_technique_id == fiche_id))

        # 4. Supprimer les actions
        session.exec(delete(ActionBudgetaire).where(ActionBudgetaire.fiche_technique_id == fiche_id))

        # 5. Supprimer la fiche
        numero_fiche = fiche.numero_fiche
        session.delete(fiche)
        session.commit()

        logger.info(f"✅ Fiche {numero_fiche} supprimée par {current_user.email}")

        # Log activité
        ActivityService.log_user_activity(
            session=session,
            user=current_user,
            action_type="delete",
            target_type="fiche_technique",
            description=f"Suppression de la fiche technique {numero_fiche}",
            target_id=fiche_id,
            icon="🗑️",
        )

        return {"ok": True, "message": "Fiche supprimée avec succès"}

    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur suppression fiche {fiche_id}: {e}")
        raise HTTPException(500, f"Erreur lors de la suppression: {e!s}")


@router.post("/api/fiches/{fiche_id}/lignes")
def api_add_ligne(
    fiche_id: int,
    activite_id: int | None = Form(None),
    nature_depense_id: int = Form(...),
    libelle: str = Form(...),
    budget_n_moins_1: float = Form(0),
    budget_demande: float = Form(...),
    justification: str | None = Form(None),
    priorite: str = Form("Normale"),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Ajouter une ligne budgétaire à une fiche"""
    fiche = session.get(FicheTechnique, fiche_id)
    if not fiche:
        raise HTTPException(404, "Fiche non trouvée")

    # Ordre = dernier + 1
    max_ordre = (
        session.exec(
            select(func.max(LigneBudgetaire.ordre)).where(LigneBudgetaire.fiche_technique_id == fiche_id)
        ).one()
        or 0
    )

    ligne = LigneBudgetaire(
        fiche_technique_id=fiche_id,
        activite_id=activite_id,
        nature_depense_id=nature_depense_id,
        libelle=libelle,
        budget_n_moins_1=Decimal(str(budget_n_moins_1)),
        budget_demande=Decimal(str(budget_demande)),
        justification=justification,
        priorite=priorite,
        ordre=max_ordre + 1,
    )

    session.add(ligne)
    session.commit()

    logger.info(f"✅ Ligne budgétaire ajoutée à fiche {fiche_id}")
    return {"ok": True, "id": ligne.id}


@router.put("/api/fiches/{fiche_id}/lignes/{ligne_id}")
def api_update_ligne(
    fiche_id: int,
    ligne_id: int,
    libelle: str | None = Form(None),
    budget_n_moins_1: float | None = Form(None),
    budget_demande: float | None = Form(None),
    budget_valide: float | None = Form(None),
    justification: str | None = Form(None),
    priorite: str | None = Form(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Modifier une ligne budgétaire"""
    ligne = session.get(LigneBudgetaire, ligne_id)
    if not ligne or ligne.fiche_technique_id != fiche_id:
        raise HTTPException(404, "Ligne non trouvée")

    if libelle:
        ligne.libelle = libelle
    if budget_n_moins_1 is not None:
        ligne.budget_n_moins_1 = Decimal(str(budget_n_moins_1))
    if budget_demande is not None:
        ligne.budget_demande = Decimal(str(budget_demande))
    if budget_valide is not None:
        ligne.budget_valide = Decimal(str(budget_valide))
    if justification:
        ligne.justification = justification
    if priorite:
        ligne.priorite = priorite

    ligne.updated_at = datetime.utcnow()
    session.add(ligne)
    session.commit()

    logger.info(f"✅ Ligne {ligne_id} mise à jour")
    return {"ok": True}


@router.delete("/api/fiches/{fiche_id}/lignes/{ligne_id}")
def api_delete_ligne(
    fiche_id: int,
    ligne_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Supprimer une ligne budgétaire"""
    ligne = session.get(LigneBudgetaire, ligne_id)
    if not ligne or ligne.fiche_technique_id != fiche_id:
        raise HTTPException(404, "Ligne non trouvée")

    ligne.actif = False
    session.add(ligne)
    session.commit()

    logger.info(f"✅ Ligne {ligne_id} supprimée")
    return {"ok": True}


# ============================================
# UPLOAD DOCUMENTS
# ============================================


@router.get("/api/fiches/{fiche_id}/documents")
def api_get_documents_fiche(
    fiche_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """Récupérer la liste des documents d'une fiche technique"""
    fiche = session.get(FicheTechnique, fiche_id)
    if not fiche:
        raise HTTPException(404, "Fiche non trouvée")

    documents = session.exec(select(DocumentBudget).where(DocumentBudget.fiche_technique_id == fiche_id)).all()

    return {
        "ok": True,
        "documents": [
            {
                "id": doc.id,
                "nom_fichier": doc.nom_fichier,
                "filename": doc.nom_fichier,
                "description": doc.type_document,
                "taille_octets": doc.taille_octets,
                "created_at": doc.uploaded_at.isoformat() if doc.uploaded_at else None,
                "date_upload": doc.uploaded_at.isoformat() if doc.uploaded_at else None,
            }
            for doc in documents
        ],
    }


@router.post("/api/fiches/{fiche_id}/documents")
async def api_upload_document(
    fiche_id: int,
    fichier: UploadFile = File(...),
    description: str | None = Form(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Uploader un document pour une fiche technique"""
    fiche = session.get(FicheTechnique, fiche_id)
    if not fiche:
        raise HTTPException(404, "Fiche non trouvée")

    # Créer le dossier avec path_config
    from app.core.path_config import path_config

    relative_path = f"budget/fiches/{fiche_id}/{fichier.filename}"
    docs_dir = path_config.UPLOADS_DIR / "budget" / "fiches" / str(fiche_id)
    docs_dir.mkdir(parents=True, exist_ok=True)

    # Sauvegarder le fichier
    file_path = docs_dir / fichier.filename
    content = await fichier.read()

    with open(file_path, "wb") as f:
        f.write(content)

    # Enregistrer en BDD avec URL générée correctement
    file_url = path_config.get_file_url("uploads", relative_path)
    doc = DocumentBudget(
        fiche_technique_id=fiche_id,
        type_document=description or "Document général",
        nom_fichier=fichier.filename,
        file_path=file_url,
        taille_octets=len(content),
        uploaded_by_user_id=current_user.id,
    )

    session.add(doc)
    session.commit()

    logger.info(f"✅ Document uploadé : {fichier.filename} pour fiche {fiche_id}")
    return {"ok": True, "id": doc.id, "filename": fichier.filename}


@router.get("/api/fiches/{fiche_id}/documents/{document_id}/download")
def api_download_document_fiche(
    fiche_id: int,
    document_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Télécharger un document d'une fiche"""
    doc = session.get(DocumentBudget, document_id)
    if not doc or doc.fiche_technique_id != fiche_id:
        raise HTTPException(404, "Document non trouvé")

    file_path = Path(f"app{doc.file_path}")
    if not file_path.exists():
        raise HTTPException(404, "Fichier physique non trouvé")

    return FileResponse(path=file_path, filename=doc.nom_fichier, media_type="application/octet-stream")


@router.delete("/api/fiches/{fiche_id}/documents/{document_id}")
def api_delete_document_fiche(
    fiche_id: int,
    document_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Supprimer un document d'une fiche"""
    doc = session.get(DocumentBudget, document_id)
    if not doc or doc.fiche_technique_id != fiche_id:
        raise HTTPException(404, "Document non trouvé")

    try:
        # Supprimer le fichier physique
        file_path = Path(f"app{doc.file_path}")
        if file_path.exists():
            file_path.unlink()
            logger.info(f"📎 Fichier supprimé : {doc.nom_fichier}")

        # Supprimer de la BDD
        session.delete(doc)
        session.commit()

        logger.info(f"✅ Document {document_id} supprimé de la fiche {fiche_id}")
        return {"ok": True, "message": "Document supprimé avec succès"}
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur suppression document {document_id}: {e}")
        raise HTTPException(500, f"Erreur lors de la suppression: {e!s}")


# ============================================
# CHANGEMENT DE STATUT
# ============================================


@router.put("/api/fiches/{fiche_id}/statut")
def api_changer_statut_fiche(
    fiche_id: int,
    nouveau_statut: str = Form(...),
    commentaire: str | None = Form(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Changer le statut d'une fiche technique"""
    fiche = session.get(FicheTechnique, fiche_id)
    if not fiche:
        raise HTTPException(404, "Fiche technique non trouvée")

    # Valider le nouveau statut
    statuts_valides = ["Brouillon", "Soumis", "Validé", "Rejeté", "Approuvé"]
    if nouveau_statut not in statuts_valides:
        raise HTTPException(400, f"Statut invalide. Valeurs possibles : {', '.join(statuts_valides)}")

    try:
        ancien_statut = fiche.statut

        # Mettre à jour le statut
        fiche.statut = nouveau_statut
        fiche.updated_by_user_id = current_user.id

        # Mettre à jour les dates selon le statut
        from datetime import date

        if nouveau_statut == "Soumis" and not fiche.date_soumission:
            fiche.date_soumission = date.today()
        elif nouveau_statut in ["Validé", "Approuvé"] and not fiche.date_validation:
            fiche.date_validation = date.today()

        session.add(fiche)

        # Créer un historique du changement de statut
        historique = HistoriqueBudget(
            fiche_technique_id=fiche_id,
            action=f"Changement de statut: {ancien_statut} → {nouveau_statut}",
            ancien_statut=ancien_statut,
            nouveau_statut=nouveau_statut,
            commentaire=commentaire,
            user_id=current_user.id,
        )
        session.add(historique)

        session.commit()

        logger.info(
            f"✅ Statut de la fiche {fiche_id} changé de '{ancien_statut}' à '{nouveau_statut}' par {current_user.email}"
        )
        return {
            "ok": True,
            "message": f"Statut changé de '{ancien_statut}' à '{nouveau_statut}'",
            "ancien_statut": ancien_statut,
            "nouveau_statut": nouveau_statut,
        }
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur changement statut fiche {fiche_id}: {e}")
        raise HTTPException(500, f"Erreur lors du changement de statut: {e!s}")


# ============================================
# IMPORT EXCEL
# ============================================


@router.post("/api/import/activites")
async def api_import_activites_excel(
    fichier: UploadFile = File(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """
    Importer des activités depuis un fichier Excel

    Format attendu:
    Code | Libelle | Programme | Direction | Nature | Description
    """
    try:
        # Lire le fichier Excel
        content = await fichier.read()
        df = pd.read_excel(io.BytesIO(content))

        # Vérifier les colonnes
        required_cols = ["Code", "Libelle"]
        if not all(col in df.columns for col in required_cols):
            raise HTTPException(400, f"Colonnes requises: {', '.join(required_cols)}")

        # Référentiels
        programmes = {p.code: p for p in session.exec(select(Programme)).all()}
        directions = {d.code: d for d in session.exec(select(Direction)).all()}
        natures = {n.code: n for n in session.exec(select(NatureDepense)).all()}

        count_created = 0
        count_updated = 0
        errors = []

        for idx, row in df.iterrows():
            try:
                code = str(row["Code"]).strip()
                libelle = str(row["Libelle"]).strip()

                # Rechercher activité existante
                existing = session.exec(select(Activite).where(Activite.code == code)).first()

                # Programme, Direction, Nature (optionnels)
                prog_id = None
                if "Programme" in row and pd.notna(row["Programme"]):
                    prog_code = str(row["Programme"]).strip()
                    if prog_code in programmes:
                        prog_id = programmes[prog_code].id

                dir_id = None
                if "Direction" in row and pd.notna(row["Direction"]):
                    dir_code = str(row["Direction"]).strip()
                    if dir_code in directions:
                        dir_id = directions[dir_code].id

                nat_id = None
                if "Nature" in row and pd.notna(row["Nature"]):
                    nat_code = str(row["Nature"]).strip()
                    if nat_code in natures:
                        nat_id = natures[nat_code].id

                desc = str(row["Description"]) if "Description" in row and pd.notna(row["Description"]) else None

                if existing:
                    # Mise à jour
                    existing.libelle = libelle
                    existing.programme_id = prog_id
                    existing.direction_id = dir_id
                    existing.nature_depense_id = nat_id
                    existing.description = desc
                    existing.updated_at = datetime.utcnow()
                    session.add(existing)
                    count_updated += 1
                else:
                    # Création
                    activite = Activite(
                        code=code,
                        libelle=libelle,
                        programme_id=prog_id,
                        direction_id=dir_id,
                        nature_depense_id=nat_id,
                        description=desc,
                    )
                    session.add(activite)
                    count_created += 1

            except Exception as e:
                errors.append(f"Ligne {idx + 2}: {e!s}")

        session.commit()

        logger.info(f"✅ Import activités : {count_created} créées, {count_updated} mises à jour")

        return {"ok": True, "created": count_created, "updated": count_updated, "errors": errors}

    except Exception as e:
        session.rollback()
        logger.error(f"Erreur import Excel: {e}")
        raise HTTPException(500, f"Erreur lors de l'import: {e!s}")


# ============================================
# EXPORT PDF
# ============================================


@router.get("/api/fiches/{fiche_id}/export/pdf")
def api_export_fiche_pdf(
    fiche_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """
    Exporter une fiche technique en PDF avec ReportLab
    Génère un vrai PDF professionnel et téléchargeable
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import cm
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
        from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

        from app.core.config import settings
    except ImportError:
        raise HTTPException(500, "ReportLab n'est pas installé. Exécutez: pip install reportlab")

    fiche = session.get(FicheTechnique, fiche_id)
    if not fiche:
        raise HTTPException(404, "Fiche technique non trouvée")

    # Récupérer la structure hiérarchique
    actions = session.exec(
        select(ActionBudgetaire).where(ActionBudgetaire.fiche_technique_id == fiche_id).order_by(ActionBudgetaire.ordre)
    ).all()

    for action in actions:
        services = session.exec(
            select(ServiceBeneficiaire)
            .where(ServiceBeneficiaire.action_id == action.id)
            .order_by(ServiceBeneficiaire.ordre)
        ).all()
        object.__setattr__(action, "services", services)

        for service in services:
            activites = session.exec(
                select(ActiviteBudgetaire)
                .where(ActiviteBudgetaire.service_beneficiaire_id == service.id)
                .order_by(ActiviteBudgetaire.ordre)
            ).all()
            object.__setattr__(service, "activites", activites)

            for activite in activites:
                lignes = session.exec(
                    select(LigneBudgetaireDetail)
                    .where(LigneBudgetaireDetail.activite_id == activite.id)
                    .order_by(LigneBudgetaireDetail.ordre)
                ).all()
                object.__setattr__(activite, "lignes", lignes)

    from collections import defaultdict

    actions_par_nature = defaultdict(list)
    for action in actions:
        nature = action.nature_depense or "BIENS ET SERVICES"
        actions_par_nature[nature].append(action)

    programme = session.get(Programme, fiche.programme_id)

    # Fonction pour ajouter en-tête et pied de page
    def add_page_number(canvas, doc):
        """Ajouter en-tête et pied de page sur chaque page"""
        canvas.saveState()

        # En-tête sur chaque page
        canvas.setFont("Helvetica-Bold", 9)
        canvas.setFillColor(colors.HexColor("#2c3e50"))
        canvas.drawCentredString(
            landscape(A4)[0] / 2,
            landscape(A4)[1] - 1.2 * cm,
            f"{fiche.numero_fiche} | Budget {fiche.annee_budget} | {programme.libelle[:50]}",
        )

        # Ligne de séparation en-tête
        canvas.setStrokeColor(colors.HexColor("#667eea"))
        canvas.setLineWidth(1)
        canvas.line(1 * cm, landscape(A4)[1] - 1.5 * cm, landscape(A4)[0] - 1 * cm, landscape(A4)[1] - 1.5 * cm)

        # Pied de page avec numéro de page
        page_num = canvas.getPageNumber()
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(colors.grey)

        # Gauche : Date et utilisateur
        canvas.drawString(
            0.5 * cm, 0.5 * cm, f"Édité le {fiche.created_at.strftime('%d/%m/%Y')} par {current_user.email}"
        )

        # Centre : Numéro de page
        canvas.drawCentredString(landscape(A4)[0] / 2, 0.5 * cm, f"Page {page_num}")

        # Droite : Application
        canvas.drawRightString(landscape(A4)[0] - 0.5 * cm, 0.5 * cm, settings.APP_NAME)

        # Ligne de séparation pied de page
        canvas.setStrokeColor(colors.HexColor("#667eea"))
        canvas.setLineWidth(1)
        canvas.line(0.5 * cm, 0.8 * cm, landscape(A4)[0] - 0.5 * cm, 0.8 * cm)

        canvas.restoreState()

    # Créer le PDF avec marges réduites
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=2.5 * cm,
        bottomMargin=2 * cm,
        title=f"Fiche Technique {fiche.numero_fiche}",
        author=current_user.email,
        subject=f"Budget {fiche.annee_budget} - {programme.libelle}",
    )

    # Styles
    styles = getSampleStyleSheet()

    # Style pour le titre
    titre_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=16,
        textColor=colors.HexColor("#2c3e50"),
        alignment=TA_CENTER,
        spaceAfter=12,
        fontName="Helvetica-Bold",
    )

    # Style pour les infos
    info_style = ParagraphStyle("Info", parent=styles["Normal"], fontSize=10, alignment=TA_CENTER, spaceAfter=6)

    # Construire le document
    elements = []

    # En-tête
    elements.append(Paragraph("📋 FICHE TECHNIQUE BUDGÉTAIRE HIÉRARCHIQUE", titre_style))
    elements.append(
        Paragraph(f"<b>{fiche.numero_fiche}</b> | Budget {fiche.annee_budget} | {fiche.statut}", info_style)
    )
    elements.append(Paragraph(f"{programme.libelle} ({programme.code})", info_style))
    elements.append(Paragraph(f"<b>Budget Total :</b> {fiche.budget_total_demande:,.0f} FCFA", info_style))
    elements.append(Spacer(1, 0.5 * cm))

    # Construire les données du tableau
    table_data = []

    # En-têtes
    headers = [
        "CODE ET LIBELLE",
        "BUDGET\nVOTÉ N\n(A)",
        "BUDGET\nACTUEL N\n(B)",
        "ENVELOPPE\nN+1\n(C)",
        "COMPLÉMENT\nSOLLICITÉ\n(D)",
        "BUDGET\nSOUHAITÉ\n(E=C+D)",
        "ENGAGEMENT\nÉTAT\n(F)",
        "AUTRE\nCOMPLÉMENT\n(G)",
        "PROJET\nBUDGET N+1\n(H=C+F+G)",
    ]
    table_data.append(headers)

    # Parcourir la hiérarchie
    for nature, actions_nature in actions_par_nature.items():
        # Header Nature avec fusion de colonnes (SPAN)
        nature_row = [nature.upper(), "", "", "", "", "", "", "", ""]
        table_data.append(nature_row)

        for action in actions_nature:
            # Action avec texte qui peut se retourner à la ligne
            action_text = Paragraph(
                action.libelle.replace("- ", "").strip(), ParagraphStyle("ActionText", fontSize=6, leftIndent=0)
            )
            table_data.append(
                [
                    action_text,
                    f"{action.budget_vote_n:,.0f}",
                    f"{action.budget_actuel_n:,.0f}",
                    f"{action.enveloppe_n_plus_1:,.0f}",
                    f"{action.complement_solicite:,.0f}",
                    f"{action.budget_souhaite:,.0f}",
                    f"{action.engagement_etat:,.0f}",
                    f"{action.autre_complement:,.0f}",
                    f"{action.projet_budget_n_plus_1:,.0f}",
                ]
            )

            for service in action.services:
                # Service avec fusion de colonnes et texte à gauche
                service_text = Paragraph(
                    f"  {service.libelle.replace('- ', '').strip()}",
                    ParagraphStyle("ServiceText", fontSize=6, leftIndent=0.2 * cm, alignment=TA_LEFT),
                )
                service_row = [service_text, "", "", "", "", "", "", "", ""]
                table_data.append(service_row)

                for activite in service.activites:
                    # Activité avec texte qui peut se retourner à la ligne
                    activite_text = Paragraph(
                        f"    {activite.libelle.replace('- ', '').strip()}",
                        ParagraphStyle("ActiviteText", fontSize=6, leftIndent=0.4 * cm),
                    )
                    table_data.append(
                        [
                            activite_text,
                            f"{activite.budget_vote_n:,.0f}",
                            f"{activite.budget_actuel_n:,.0f}",
                            f"{activite.enveloppe_n_plus_1:,.0f}",
                            f"{activite.complement_solicite:,.0f}",
                            f"{activite.budget_souhaite:,.0f}",
                            f"{activite.engagement_etat:,.0f}",
                            f"{activite.autre_complement:,.0f}",
                            f"{activite.projet_budget_n_plus_1:,.0f}",
                        ]
                    )

                    # Lignes budgétaires avec texte qui peut se retourner à la ligne
                    for ligne in activite.lignes:
                        ligne_text = Paragraph(
                            f"      {ligne.code} - {ligne.libelle}",
                            ParagraphStyle("LigneText", fontSize=6, leftIndent=0.6 * cm),
                        )
                        table_data.append(
                            [
                                ligne_text,
                                f"{ligne.budget_vote_n:,.0f}",
                                f"{ligne.budget_actuel_n:,.0f}",
                                f"{ligne.enveloppe_n_plus_1:,.0f}",
                                f"{ligne.complement_solicite:,.0f}",
                                f"{ligne.budget_souhaite:,.0f}",
                                f"{ligne.engagement_etat:,.0f}",
                                f"{ligne.autre_complement:,.0f}",
                                f"{ligne.projet_budget_n_plus_1:,.0f}",
                            ]
                        )

                    # Sous-total Activité avec police adaptée aux grandes valeurs
                    sous_total_activite = [
                        "SOUS-TOTAL ACTIVITÉ",
                        f"{activite.budget_vote_n:,.0f}",
                        f"{activite.budget_actuel_n:,.0f}",
                        f"{activite.enveloppe_n_plus_1:,.0f}",
                        f"{activite.complement_solicite:,.0f}",
                        f"{activite.budget_souhaite:,.0f}",
                        f"{activite.engagement_etat:,.0f}",
                        f"{activite.autre_complement:,.0f}",
                        f"{activite.projet_budget_n_plus_1:,.0f}",
                    ]
                    table_data.append(sous_total_activite)

                # Sous-total Service avec police adaptée aux grandes valeurs
                service_totals = {
                    "vote": sum(a.budget_vote_n for a in service.activites),
                    "actuel": sum(a.budget_actuel_n for a in service.activites),
                    "enveloppe": sum(a.enveloppe_n_plus_1 for a in service.activites),
                    "complement": sum(a.complement_solicite for a in service.activites),
                    "souhaite": sum(a.budget_souhaite for a in service.activites),
                    "engagement": sum(a.engagement_etat for a in service.activites),
                    "autre": sum(a.autre_complement for a in service.activites),
                    "projet": sum(a.projet_budget_n_plus_1 for a in service.activites),
                }
                sous_total_service = [
                    "SOUS-TOTAL SERVICE",
                    f"{service_totals['vote']:,.0f}",
                    f"{service_totals['actuel']:,.0f}",
                    f"{service_totals['enveloppe']:,.0f}",
                    f"{service_totals['complement']:,.0f}",
                    f"{service_totals['souhaite']:,.0f}",
                    f"{service_totals['engagement']:,.0f}",
                    f"{service_totals['autre']:,.0f}",
                    f"{service_totals['projet']:,.0f}",
                ]
                table_data.append(sous_total_service)

            # Sous-total Action
            table_data.append(
                [
                    "SOUS-TOTAL ACTION",
                    f"{action.budget_vote_n:,.0f}",
                    f"{action.budget_actuel_n:,.0f}",
                    f"{action.enveloppe_n_plus_1:,.0f}",
                    f"{action.complement_solicite:,.0f}",
                    f"{action.budget_souhaite:,.0f}",
                    f"{action.engagement_etat:,.0f}",
                    f"{action.autre_complement:,.0f}",
                    f"{action.projet_budget_n_plus_1:,.0f}",
                ]
            )

        # Total Nature avec police adaptée aux grandes valeurs
        nature_totals = {
            "vote": sum(a.budget_vote_n for a in actions_nature),
            "actuel": sum(a.budget_actuel_n for a in actions_nature),
            "enveloppe": sum(a.enveloppe_n_plus_1 for a in actions_nature),
            "complement": sum(a.complement_solicite for a in actions_nature),
            "souhaite": sum(a.budget_souhaite for a in actions_nature),
            "engagement": sum(a.engagement_etat for a in actions_nature),
            "autre": sum(a.autre_complement for a in actions_nature),
            "projet": sum(a.projet_budget_n_plus_1 for a in actions_nature),
        }
        total_nature = [
            f"TOTAL {nature.upper()}",
            f"{nature_totals['vote']:,.0f}",
            f"{nature_totals['actuel']:,.0f}",
            f"{nature_totals['enveloppe']:,.0f}",
            f"{nature_totals['complement']:,.0f}",
            f"{nature_totals['souhaite']:,.0f}",
            f"{nature_totals['engagement']:,.0f}",
            f"{nature_totals['autre']:,.0f}",
            f"{nature_totals['projet']:,.0f}",
        ]
        table_data.append(total_nature)

    # Total Général avec police adaptée aux grandes valeurs
    total_general = {
        "vote": sum(sum(a.budget_vote_n for a in acts) for acts in actions_par_nature.values()),
        "actuel": sum(sum(a.budget_actuel_n for a in acts) for acts in actions_par_nature.values()),
        "enveloppe": sum(sum(a.enveloppe_n_plus_1 for a in acts) for acts in actions_par_nature.values()),
        "complement": sum(sum(a.complement_solicite for a in acts) for acts in actions_par_nature.values()),
        "souhaite": sum(sum(a.budget_souhaite for a in acts) for acts in actions_par_nature.values()),
        "engagement": sum(sum(a.engagement_etat for a in acts) for acts in actions_par_nature.values()),
        "autre": sum(sum(a.autre_complement for a in acts) for acts in actions_par_nature.values()),
        "projet": sum(sum(a.projet_budget_n_plus_1 for a in acts) for acts in actions_par_nature.values()),
    }
    total_general_row = [
        "TOTAL GÉNÉRAL",
        f"{total_general['vote']:,.0f}",
        f"{total_general['actuel']:,.0f}",
        f"{total_general['enveloppe']:,.0f}",
        f"{total_general['complement']:,.0f}",
        f"{total_general['souhaite']:,.0f}",
        f"{total_general['engagement']:,.0f}",
        f"{total_general['autre']:,.0f}",
        f"{total_general['projet']:,.0f}",
    ]
    table_data.append(total_general_row)

    # Créer le tableau avec largeurs optimisées - colonne texte réduite
    col_widths = [7 * cm, 2.8 * cm, 2.8 * cm, 2.8 * cm, 2.8 * cm, 2.8 * cm, 2.8 * cm, 2.8 * cm, 3 * cm]
    table = Table(table_data, colWidths=col_widths, repeatRows=1)

    # Styles du tableau
    table_style = TableStyle(
        [
            # En-têtes
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#ef8d4b")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("ALIGN", (0, 0), (-1, 0), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 7),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
            ("TOPPADDING", (0, 0), (-1, 0), 8),
            # Toutes les colonnes centrées horizontalement
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            # Bordures
            ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            # Police générale avec taille réduite
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 6),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),  # Centrage vertical pour toutes les cellules
            # Padding réduit pour économiser l'espace
            ("LEFTPADDING", (0, 0), (-1, -1), 2),
            ("RIGHTPADDING", (0, 0), (-1, -1), 2),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ]
    )

    # Appliquer les couleurs selon le type de ligne
    row_index = 1  # Commencer après les headers
    for nature, actions_nature in actions_par_nature.items():
        # Header Nature (jaune or) avec fusion de colonnes
        table_style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#ffd700"))
        table_style.add("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold")
        table_style.add("FONTSIZE", (0, row_index), (-1, row_index), 8)
        table_style.add("ALIGN", (0, row_index), (-1, row_index), "CENTER")
        table_style.add("SPAN", (0, row_index), (-1, row_index))  # Fusionner toutes les colonnes
        row_index += 1

        for action in actions_nature:
            # Action (bleu)
            table_style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#9bc2e6"))
            table_style.add("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold")
            table_style.add("FONTSIZE", (0, row_index), (-1, row_index), 8)
            row_index += 1

            for service in action.services:
                # Service (jaune) avec fusion de colonnes et texte à gauche
                table_style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#ffc000"))
                table_style.add("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold")
                table_style.add("FONTSIZE", (0, row_index), (-1, row_index), 6)  # Police hiérarchique
                table_style.add("ALIGN", (0, row_index), (-1, row_index), "LEFT")  # Texte à gauche
                table_style.add("SPAN", (0, row_index), (-1, row_index))  # Fusionner toutes les colonnes
                row_index += 1

                for activite in service.activites:
                    # Activité (vert)
                    table_style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#92d050"))
                    table_style.add("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold")
                    row_index += 1

                    # Lignes (blanc)
                    for ligne in activite.lignes:
                        table_style.add("FONTSIZE", (0, row_index), (0, row_index), 6)
                        row_index += 1

                    # Sous-total Activité (vert clair) avec police hiérarchique
                    table_style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#d4edda"))
                    table_style.add("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold")
                    table_style.add("FONTSIZE", (0, row_index), (-1, row_index), 6)  # Police hiérarchique
                    table_style.add("ALIGN", (0, row_index), (0, row_index), "RIGHT")
                    row_index += 1

                # Sous-total Service (jaune clair) avec police hiérarchique
                table_style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#fff3cd"))
                table_style.add("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold")
                table_style.add("FONTSIZE", (0, row_index), (-1, row_index), 7)  # Police hiérarchique
                table_style.add("ALIGN", (0, row_index), (0, row_index), "RIGHT")
                row_index += 1

            # Sous-total Action (bleu clair) avec police hiérarchique
            table_style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#cce5ff"))
            table_style.add("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold")
            table_style.add("FONTSIZE", (0, row_index), (-1, row_index), 8)  # Police hiérarchique
            table_style.add("ALIGN", (0, row_index), (0, row_index), "RIGHT")
            row_index += 1

        # Total Nature (or) avec police hiérarchique
        table_style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#ffd700"))
        table_style.add("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold")
        table_style.add("FONTSIZE", (0, row_index), (-1, row_index), 9)  # Police hiérarchique
        table_style.add("ALIGN", (0, row_index), (0, row_index), "RIGHT")
        row_index += 1

    # Total Général (rouge) avec police hiérarchique maximale
    table_style.add("BACKGROUND", (0, row_index), (-1, row_index), colors.HexColor("#dc3545"))
    table_style.add("TEXTCOLOR", (0, row_index), (-1, row_index), colors.white)
    table_style.add("FONTNAME", (0, row_index), (-1, row_index), "Helvetica-Bold")
    table_style.add("FONTSIZE", (0, row_index), (-1, row_index), 10)  # Police hiérarchique maximale
    table_style.add("ALIGN", (0, row_index), (0, row_index), "RIGHT")

    table.setStyle(table_style)
    elements.append(table)

    # Footer
    elements.append(Spacer(1, 0.5 * cm))
    footer_style = ParagraphStyle(
        "Footer", parent=styles["Normal"], fontSize=8, alignment=TA_RIGHT, textColor=colors.grey
    )
    elements.append(Paragraph(f"Document édité le : {fiche.created_at.strftime('%d/%m/%Y à %H:%M')}", footer_style))
    elements.append(Paragraph(f"Phase : {fiche.phase} | Statut : {fiche.statut}", footer_style))
    elements.append(Paragraph(settings.APP_NAME, footer_style))

    # Générer le PDF avec en-tête et pied de page
    doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)

    # Retourner le PDF
    buffer.seek(0)

    logger.info(f"✅ PDF ReportLab généré pour fiche {fiche.numero_fiche}")

    # Nettoyer le nom du programme pour le nom de fichier
    programme_nom_clean = (
        programme.libelle.replace(" ", "_")
        .replace("/", "_")
        .replace("\\", "_")
        .replace(":", "_")
        .replace("*", "_")
        .replace("?", "_")
        .replace('"', "_")
        .replace("<", "_")
        .replace(">", "_")
        .replace("|", "_")
    )

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=Fiche_{programme_nom_clean}_{fiche.annee_budget}_{fiche.numero_fiche}.pdf"
        },
    )


@router.get("/api/fiches/{fiche_id}/export/excel")
def api_export_fiche_excel(
    fiche_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """
    Exporter une fiche technique en Excel avec la structure hiérarchique et les couleurs
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl n'est pas installé")

    fiche = session.get(FicheTechnique, fiche_id)
    if not fiche:
        raise HTTPException(404, "Fiche technique non trouvée")

    # Récupérer la structure hiérarchique
    actions = session.exec(
        select(ActionBudgetaire).where(ActionBudgetaire.fiche_technique_id == fiche_id).order_by(ActionBudgetaire.ordre)
    ).all()

    for action in actions:
        services = session.exec(
            select(ServiceBeneficiaire)
            .where(ServiceBeneficiaire.action_id == action.id)
            .order_by(ServiceBeneficiaire.ordre)
        ).all()
        object.__setattr__(action, "services", services)

        for service in services:
            activites = session.exec(
                select(ActiviteBudgetaire)
                .where(ActiviteBudgetaire.service_beneficiaire_id == service.id)
                .order_by(ActiviteBudgetaire.ordre)
            ).all()
            object.__setattr__(service, "activites", activites)

            for activite in activites:
                lignes = session.exec(
                    select(LigneBudgetaireDetail)
                    .where(LigneBudgetaireDetail.activite_id == activite.id)
                    .order_by(LigneBudgetaireDetail.ordre)
                ).all()
                object.__setattr__(activite, "lignes", lignes)

    from collections import defaultdict

    actions_par_nature = defaultdict(list)
    for action in actions:
        nature = action.nature_depense or "BIENS ET SERVICES"
        actions_par_nature[nature].append(action)

    programme = session.get(Programme, fiche.programme_id)

    # Créer le classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Fiche {fiche.numero_fiche}"

    # Styles
    header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border_style = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Couleurs hiérarchiques
    nature_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    action_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")
    service_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
    activite_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
    ligne_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    sous_total_activite_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
    sous_total_service_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
    sous_total_action_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    total_nature_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    total_general_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")

    # En-têtes
    headers = [
        "CODE ET LIBELLE",
        "BUDGET VOTÉ N",
        "BUDGET ACTUEL N",
        "ENVELOPPE N+1",
        "COMPLEMENT SOLLICITÉ",
        "BUDGET SOUHAITÉ",
        "ENGAGEMENT ÉTAT",
        "AUTRE COMPLEMENT",
        "PROJET BUDGET N+1",
        "JUSTIFICATIFS",
    ]

    current_row = 1
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_style

    # Ajuster largeurs
    ws.column_dimensions["A"].width = 50
    for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["J"].width = 40

    current_row = 2

    # Parcourir la hiérarchie
    for nature, actions_nature in actions_par_nature.items():
        # Nature
        cell = ws.cell(row=current_row, column=1)
        cell.value = nature.upper()
        cell.fill = nature_fill
        cell.font = Font(bold=True, size=11)
        cell.border = border_style
        cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(2, 11):
            ws.cell(row=current_row, column=col).border = border_style
            ws.cell(row=current_row, column=col).fill = nature_fill
        current_row += 1

        for action in actions_nature:
            # Action
            row_data = [
                f"Action : {action.libelle}",
                float(action.budget_vote_n),
                float(action.budget_actuel_n),
                float(action.enveloppe_n_plus_1),
                float(action.complement_solicite),
                float(action.budget_souhaite),
                float(action.engagement_etat),
                float(action.autre_complement),
                float(action.projet_budget_n_plus_1),
                action.justificatifs or "",
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.fill = action_fill
                cell.font = Font(bold=True, size=10)
                cell.border = border_style
                if col_num > 1 and col_num < 10:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            current_row += 1

            for service in action.services:
                # Service
                cell = ws.cell(row=current_row, column=1)
                cell.value = f"Service Bénéficiaire : {service.libelle}"
                cell.fill = service_fill
                cell.font = Font(bold=True, size=9)
                cell.border = border_style
                cell.alignment = Alignment(horizontal="left", vertical="center")
                for col in range(2, 11):
                    ws.cell(row=current_row, column=col).border = border_style
                    ws.cell(row=current_row, column=col).fill = service_fill
                current_row += 1

                for activite in service.activites:
                    # Activité
                    row_data = [
                        f"Activité : {activite.libelle}",
                        float(activite.budget_vote_n),
                        float(activite.budget_actuel_n),
                        float(activite.enveloppe_n_plus_1),
                        float(activite.complement_solicite),
                        float(activite.budget_souhaite),
                        float(activite.engagement_etat),
                        float(activite.autre_complement),
                        float(activite.projet_budget_n_plus_1),
                        activite.justificatifs or "",
                    ]
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.value = value
                        cell.fill = activite_fill
                        cell.font = Font(bold=True, size=9)
                        cell.border = border_style
                        if col_num > 1 and col_num < 10:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = "#,##0"
                        else:
                            cell.alignment = Alignment(horizontal="left", vertical="center")
                    current_row += 1

                    # Lignes budgétaires
                    for ligne in activite.lignes:
                        row_data = [
                            f"{ligne.code} - {ligne.libelle}",
                            float(ligne.budget_vote_n),
                            float(ligne.budget_actuel_n),
                            float(ligne.enveloppe_n_plus_1),
                            float(ligne.complement_solicite),
                            float(ligne.budget_souhaite),
                            float(ligne.engagement_etat),
                            float(ligne.autre_complement),
                            float(ligne.projet_budget_n_plus_1),
                            ligne.justificatifs or "",
                        ]
                        for col_num, value in enumerate(row_data, 1):
                            cell = ws.cell(row=current_row, column=col_num)
                            cell.value = value
                            cell.fill = ligne_fill
                            cell.font = Font(size=9)
                            cell.border = border_style
                            if col_num > 1 and col_num < 10:
                                cell.alignment = Alignment(horizontal="right", vertical="center")
                                cell.number_format = "#,##0"
                            else:
                                cell.alignment = Alignment(horizontal="left", vertical="center")
                        current_row += 1

                    # Sous-total Activité
                    row_data = [
                        f"SOUS-TOTAL Activité : {activite.libelle}",
                        float(activite.budget_vote_n),
                        float(activite.budget_actuel_n),
                        float(activite.enveloppe_n_plus_1),
                        float(activite.complement_solicite),
                        float(activite.budget_souhaite),
                        float(activite.engagement_etat),
                        float(activite.autre_complement),
                        float(activite.projet_budget_n_plus_1),
                        "",
                    ]
                    for col_num, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col_num)
                        cell.value = value
                        cell.fill = sous_total_activite_fill
                        cell.font = Font(bold=True, size=8, italic=True)
                        cell.border = border_style
                        if col_num > 1 and col_num < 10:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                            cell.number_format = "#,##0"
                        else:
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                    current_row += 1

                # Sous-total Service
                service_totals = {
                    "vote": sum(a.budget_vote_n for a in service.activites),
                    "actuel": sum(a.budget_actuel_n for a in service.activites),
                    "enveloppe": sum(a.enveloppe_n_plus_1 for a in service.activites),
                    "complement": sum(a.complement_solicite for a in service.activites),
                    "souhaite": sum(a.budget_souhaite for a in service.activites),
                    "engagement": sum(a.engagement_etat for a in service.activites),
                    "autre": sum(a.autre_complement for a in service.activites),
                    "projet": sum(a.projet_budget_n_plus_1 for a in service.activites),
                }
                row_data = [
                    f"SOUS-TOTAL Service : {service.libelle}",
                    float(service_totals["vote"]),
                    float(service_totals["actuel"]),
                    float(service_totals["enveloppe"]),
                    float(service_totals["complement"]),
                    float(service_totals["souhaite"]),
                    float(service_totals["engagement"]),
                    float(service_totals["autre"]),
                    float(service_totals["projet"]),
                    "",
                ]
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = value
                    cell.fill = sous_total_service_fill
                    cell.font = Font(bold=True, size=9, italic=True)
                    cell.border = border_style
                    if col_num > 1 and col_num < 10:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "#,##0"
                    else:
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                current_row += 1

            # Sous-total Action
            row_data = [
                f"SOUS-TOTAL Action : {action.libelle}",
                float(action.budget_vote_n),
                float(action.budget_actuel_n),
                float(action.enveloppe_n_plus_1),
                float(action.complement_solicite),
                float(action.budget_souhaite),
                float(action.engagement_etat),
                float(action.autre_complement),
                float(action.projet_budget_n_plus_1),
                "",
            ]
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.fill = sous_total_action_fill
                cell.font = Font(bold=True, size=10, italic=True)
                cell.border = border_style
                if col_num > 1 and col_num < 10:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            current_row += 1

        # Total Nature
        nature_totals = {
            "vote": sum(a.budget_vote_n for a in actions_nature),
            "actuel": sum(a.budget_actuel_n for a in actions_nature),
            "enveloppe": sum(a.enveloppe_n_plus_1 for a in actions_nature),
            "complement": sum(a.complement_solicite for a in actions_nature),
            "souhaite": sum(a.budget_souhaite for a in actions_nature),
            "engagement": sum(a.engagement_etat for a in actions_nature),
            "autre": sum(a.autre_complement for a in actions_nature),
            "projet": sum(a.projet_budget_n_plus_1 for a in actions_nature),
        }
        row_data = [
            f"TOTAL {nature.upper()}",
            float(nature_totals["vote"]),
            float(nature_totals["actuel"]),
            float(nature_totals["enveloppe"]),
            float(nature_totals["complement"]),
            float(nature_totals["souhaite"]),
            float(nature_totals["engagement"]),
            float(nature_totals["autre"]),
            float(nature_totals["projet"]),
            "",
        ]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = value
            cell.fill = total_nature_fill
            cell.font = Font(bold=True, size=11)
            cell.border = border_style
            if col_num > 1 and col_num < 10:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
        current_row += 1

        # Ligne vide entre natures
        current_row += 1

    # Total Général
    total_general = {
        "vote": sum(sum(a.budget_vote_n for a in acts) for acts in actions_par_nature.values()),
        "actuel": sum(sum(a.budget_actuel_n for a in acts) for acts in actions_par_nature.values()),
        "enveloppe": sum(sum(a.enveloppe_n_plus_1 for a in acts) for acts in actions_par_nature.values()),
        "complement": sum(sum(a.complement_solicite for a in acts) for acts in actions_par_nature.values()),
        "souhaite": sum(sum(a.budget_souhaite for a in acts) for acts in actions_par_nature.values()),
        "engagement": sum(sum(a.engagement_etat for a in acts) for acts in actions_par_nature.values()),
        "autre": sum(sum(a.autre_complement for a in acts) for acts in actions_par_nature.values()),
        "projet": sum(sum(a.projet_budget_n_plus_1 for a in acts) for acts in actions_par_nature.values()),
    }
    row_data = [
        "TOTAL GÉNÉRAL",
        float(total_general["vote"]),
        float(total_general["actuel"]),
        float(total_general["enveloppe"]),
        float(total_general["complement"]),
        float(total_general["souhaite"]),
        float(total_general["engagement"]),
        float(total_general["autre"]),
        float(total_general["projet"]),
        "",
    ]
    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = value
        cell.fill = total_general_fill
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.border = border_style
        if col_num > 1 and col_num < 10:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "#,##0"
        else:
            cell.alignment = Alignment(horizontal="right", vertical="center")

    # Figer les en-têtes
    ws.freeze_panes = "A2"

    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(f"✅ Excel généré pour fiche {fiche.numero_fiche}")

    # Nom du fichier
    programme_nom_clean = programme.libelle.replace(" ", "_").replace("/", "_")

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=Fiche_{programme_nom_clean}_{fiche.annee_budget}_{fiche.numero_fiche}.xlsx"
        },
    )


# ============================================
# FICHES TECHNIQUES HIÉRARCHIQUES
# ============================================


@router.get("/fiches/{fiche_id}/structure", response_class=HTMLResponse, name="budget_fiche_structure")
def budget_fiche_structure(
    fiche_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Afficher la structure hiérarchique d'une fiche technique"""
    fiche = session.get(FicheTechnique, fiche_id)
    if not fiche:
        raise HTTPException(404, "Fiche technique non trouvée")

    # Recalculer les totaux pour s'assurer que tout est à jour
    FicheTechniqueService._recalculer_totaux_hierarchie(fiche_id, session)

    # Recharger la fiche après recalcul
    session.expire(fiche)
    fiche = session.get(FicheTechnique, fiche_id)

    # Récupérer la structure hiérarchique
    actions = session.exec(
        select(ActionBudgetaire).where(ActionBudgetaire.fiche_technique_id == fiche_id).order_by(ActionBudgetaire.ordre)
    ).all()

    # Charger les services pour chaque action
    for action in actions:
        services = session.exec(
            select(ServiceBeneficiaire)
            .where(ServiceBeneficiaire.action_id == action.id)
            .order_by(ServiceBeneficiaire.ordre)
        ).all()
        object.__setattr__(action, "services", services)

        # Charger les activités pour chaque service
        for service in services:
            activites = session.exec(
                select(ActiviteBudgetaire)
                .where(ActiviteBudgetaire.service_beneficiaire_id == service.id)
                .order_by(ActiviteBudgetaire.ordre)
            ).all()
            object.__setattr__(service, "activites", activites)

            # Charger les lignes pour chaque activité
            for activite in activites:
                lignes = session.exec(
                    select(LigneBudgetaireDetail)
                    .where(LigneBudgetaireDetail.activite_id == activite.id)
                    .order_by(LigneBudgetaireDetail.ordre)
                ).all()
                object.__setattr__(activite, "lignes", lignes)

    # Grouper les actions par nature de dépense
    from collections import defaultdict

    actions_par_nature = defaultdict(list)
    for action in actions:
        nature = action.nature_depense or "BIENS ET SERVICES"
        actions_par_nature[nature].append(action)

    # Référentiels
    programme = session.get(Programme, fiche.programme_id)
    direction = session.get(Direction, fiche.direction_id) if fiche.direction_id else None

    return templates.TemplateResponse(
        "pages/budget_fiche_structure.html",
        get_template_context(
            request,
            fiche=fiche,
            actions=actions,
            actions_par_nature=dict(actions_par_nature),
            programme=programme,
            direction=direction,
            current_user=current_user,
        ),
    )


@router.get("/api/telecharger-template-fiche")
async def api_telecharger_template_fiche(annee: int = 2025, current_user: User = Depends(get_current_user)):
    """
    Télécharger un modèle Excel vierge pour créer une fiche technique
    """
    try:
        # Créer un classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Fiche Technique N"

        # Définir les styles
        header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")  # Orange
        header_font = Font(bold=True, color="FFFFFF", size=11)
        example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        border_style = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Définir les en-têtes de colonnes (notation N et N+1)
        headers = [
            "CODE / LIBELLE",
            "BUDGET VOTÉ N",
            "BUDGET ACTUEL N",
            "ENVELOPPE N+1",
            "COMPLEMENT SOLLICITÉ",
            "BUDGET SOUHAITÉ",
            "ENGAGEMENT DE L'ETAT",
            "AUTRE COMPLEMENT",
            "PROJET DE BUDGET N+1",
            "JUSTIFICATIFS",
        ]

        # Écrire les en-têtes
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_style

        # Ajuster la largeur des colonnes
        ws.column_dimensions["A"].width = 50
        for col in ["B", "C", "D", "E", "F", "G", "H", "I"]:
            ws.column_dimensions[col].width = 18
        ws.column_dimensions["J"].width = 40

        # Ajouter des exemples de structure hiérarchique complète et réaliste
        exemples = [
            # === NATURE 1 : BIENS ET SERVICES ===
            ("BIENS ET SERVICES", "", "", "", "", "", "", "", "", "Nature de dépense"),
            # Action 1.1
            (
                "Action : Pilotage et Gouvernance Institutionnelle",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "Coordination générale et gouvernance",
            ),
            # Service 1.1.1
            ("Service Bénéficiaire : Direction Générale", "", "", "", "", "", "", "", "", ""),
            # Activité 1.1.1.1
            ("Activité : Coordination Administrative", "", "", "", "", "", "", "", "", ""),
            (
                "601100 - Fournitures de bureau",
                "500000",
                "480000",
                "550000",
                "20000",
                "570000",
                "30000",
                "0",
                "580000",
                "Papeterie, consommables",
            ),
            (
                "601200 - Documentation et abonnements",
                "300000",
                "290000",
                "350000",
                "15000",
                "365000",
                "20000",
                "0",
                "370000",
                "Revues, documentation technique",
            ),
            (
                "SOUS-TOTAL Activité : Coordination Administrative",
                "800000",
                "770000",
                "900000",
                "35000",
                "935000",
                "50000",
                "0",
                "950000",
                "",
            ),
            # Activité 1.1.1.2
            ("Activité : Communication Institutionnelle", "", "", "", "", "", "", "", "", ""),
            (
                "606300 - Frais de publicité et communication",
                "800000",
                "750000",
                "900000",
                "40000",
                "940000",
                "50000",
                "0",
                "950000",
                "Campagnes de communication",
            ),
            (
                "606400 - Réception et manifestations",
                "600000",
                "580000",
                "700000",
                "30000",
                "730000",
                "40000",
                "0",
                "740000",
                "Cérémonies officielles",
            ),
            (
                "SOUS-TOTAL Activité : Communication Institutionnelle",
                "1400000",
                "1330000",
                "1600000",
                "70000",
                "1670000",
                "90000",
                "0",
                "1690000",
                "",
            ),
            (
                "SOUS-TOTAL Service : Direction Générale",
                "2200000",
                "2100000",
                "2500000",
                "105000",
                "2605000",
                "140000",
                "0",
                "2640000",
                "",
            ),
            # Service 1.1.2
            ("Service Bénéficiaire : Direction des Affaires Financières", "", "", "", "", "", "", "", "", ""),
            # Activité 1.1.2.1
            ("Activité : Gestion Budgétaire et Comptable", "", "", "", "", "", "", "", "", ""),
            (
                "601800 - Fournitures et matériel informatique",
                "1200000",
                "1150000",
                "1400000",
                "60000",
                "1460000",
                "80000",
                "0",
                "1480000",
                "Matériel bureautique",
            ),
            (
                "606100 - Frais de formation du personnel",
                "900000",
                "850000",
                "1000000",
                "45000",
                "1045000",
                "60000",
                "0",
                "1060000",
                "Formations comptables et financières",
            ),
            (
                "SOUS-TOTAL Activité : Gestion Budgétaire et Comptable",
                "2100000",
                "2000000",
                "2400000",
                "105000",
                "2505000",
                "140000",
                "0",
                "2540000",
                "",
            ),
            # Activité 1.1.2.2
            ("Activité : Suivi et Contrôle de Gestion", "", "", "", "", "", "", "", "", ""),
            (
                "602200 - Services extérieurs (audit, conseil)",
                "1500000",
                "1400000",
                "1700000",
                "75000",
                "1775000",
                "100000",
                "0",
                "1800000",
                "Audits externes, consultants",
            ),
            (
                "605000 - Logiciels et licences",
                "800000",
                "770000",
                "900000",
                "40000",
                "940000",
                "50000",
                "0",
                "950000",
                "Logiciels de gestion",
            ),
            (
                "SOUS-TOTAL Activité : Suivi et Contrôle de Gestion",
                "2300000",
                "2170000",
                "2600000",
                "115000",
                "2715000",
                "150000",
                "0",
                "2750000",
                "",
            ),
            (
                "SOUS-TOTAL Service : Direction des Affaires Financières",
                "4400000",
                "4170000",
                "5000000",
                "220000",
                "5220000",
                "290000",
                "0",
                "5290000",
                "",
            ),
            (
                "SOUS-TOTAL Action : Pilotage et Gouvernance Institutionnelle",
                "6600000",
                "6270000",
                "7500000",
                "325000",
                "7825000",
                "430000",
                "0",
                "7930000",
                "",
            ),
            # Action 1.2
            (
                "Action : Gestion des Ressources et Services Généraux",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "Gestion administrative et logistique",
            ),
            # Service 1.2.1
            ("Service Bénéficiaire : Service des Moyens Généraux", "", "", "", "", "", "", "", "", ""),
            # Activité 1.2.1.1
            ("Activité : Entretien et Maintenance des Locaux", "", "", "", "", "", "", "", "", ""),
            (
                "605200 - Entretien bâtiments",
                "2000000",
                "1900000",
                "2300000",
                "100000",
                "2400000",
                "130000",
                "0",
                "2430000",
                "Réparations, peinture",
            ),
            (
                "605300 - Maintenance des équipements",
                "1500000",
                "1450000",
                "1700000",
                "75000",
                "1775000",
                "95000",
                "0",
                "1795000",
                "Climatisation, électricité",
            ),
            (
                "SOUS-TOTAL Activité : Entretien et Maintenance des Locaux",
                "3500000",
                "3350000",
                "4000000",
                "175000",
                "4175000",
                "225000",
                "0",
                "4225000",
                "",
            ),
            # Activité 1.2.1.2
            ("Activité : Gestion du Parc Automobile", "", "", "", "", "", "", "", "", ""),
            (
                "606100 - Carburants et lubrifiants",
                "3000000",
                "2850000",
                "3400000",
                "150000",
                "3550000",
                "190000",
                "0",
                "3590000",
                "Essence, diesel",
            ),
            (
                "605400 - Réparations véhicules",
                "1800000",
                "1700000",
                "2000000",
                "90000",
                "2090000",
                "115000",
                "0",
                "2115000",
                "Entretien, pièces détachées",
            ),
            (
                "SOUS-TOTAL Activité : Gestion du Parc Automobile",
                "4800000",
                "4550000",
                "5400000",
                "240000",
                "5640000",
                "305000",
                "0",
                "5705000",
                "",
            ),
            (
                "SOUS-TOTAL Service : Service des Moyens Généraux",
                "8300000",
                "7900000",
                "9400000",
                "415000",
                "9815000",
                "530000",
                "0",
                "9930000",
                "",
            ),
            # Service 1.2.2
            ("Service Bénéficiaire : Service Informatique", "", "", "", "", "", "", "", "", ""),
            # Activité 1.2.2.1
            ("Activité : Maintenance Infrastructure IT", "", "", "", "", "", "", "", "", ""),
            (
                "601800 - Matériel réseau et serveurs",
                "2500000",
                "2400000",
                "2800000",
                "120000",
                "2920000",
                "160000",
                "0",
                "2960000",
                "Switches, routeurs, serveurs",
            ),
            (
                "605000 - Licences logicielles",
                "1200000",
                "1150000",
                "1400000",
                "60000",
                "1460000",
                "80000",
                "0",
                "1480000",
                "Windows, Office, antivirus",
            ),
            (
                "SOUS-TOTAL Activité : Maintenance Infrastructure IT",
                "3700000",
                "3550000",
                "4200000",
                "180000",
                "4380000",
                "240000",
                "0",
                "4440000",
                "",
            ),
            # Activité 1.2.2.2
            ("Activité : Support Utilisateurs et Formation", "", "", "", "", "", "", "", "", ""),
            (
                "606100 - Formation informatique",
                "800000",
                "770000",
                "900000",
                "40000",
                "940000",
                "50000",
                "0",
                "950000",
                "Formation bureautique, cybersécurité",
            ),
            (
                "622800 - Prestations de services IT",
                "1500000",
                "1450000",
                "1700000",
                "75000",
                "1775000",
                "95000",
                "0",
                "1795000",
                "Support technique externe",
            ),
            (
                "SOUS-TOTAL Activité : Support Utilisateurs et Formation",
                "2300000",
                "2220000",
                "2600000",
                "115000",
                "2715000",
                "145000",
                "0",
                "2745000",
                "",
            ),
            (
                "SOUS-TOTAL Service : Service Informatique",
                "6000000",
                "5770000",
                "6800000",
                "295000",
                "7095000",
                "385000",
                "0",
                "7185000",
                "",
            ),
            (
                "SOUS-TOTAL Action : Gestion des Ressources et Services Généraux",
                "14300000",
                "13670000",
                "16200000",
                "710000",
                "16910000",
                "915000",
                "0",
                "17115000",
                "",
            ),
            (
                "TOTAL BIENS ET SERVICES",
                "20900000",
                "19940000",
                "23700000",
                "1035000",
                "24735000",
                "1345000",
                "0",
                "25045000",
                "",
            ),
            # === NATURE 2 : PERSONNEL ===
            ("", "", "", "", "", "", "", "", "", ""),
            ("PERSONNEL", "", "", "", "", "", "", "", "", "Nature de dépense"),
            # Action 2.1
            (
                "Action : Gestion des Ressources Humaines",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "Recrutement, formation, administration du personnel",
            ),
            # Service 2.1.1
            ("Service Bénéficiaire : Direction des Ressources Humaines", "", "", "", "", "", "", "", "", ""),
            # Activité 2.1.1.1
            ("Activité : Recrutement et Mobilité", "", "", "", "", "", "", "", "", ""),
            (
                "661100 - Salaires et traitements nouveaux agents",
                "15000000",
                "14500000",
                "17000000",
                "750000",
                "17750000",
                "1000000",
                "0",
                "18000000",
                "10 nouveaux recrutements",
            ),
            (
                "661200 - Indemnités et primes",
                "3000000",
                "2900000",
                "3500000",
                "150000",
                "3650000",
                "200000",
                "0",
                "3700000",
                "Primes de rendement",
            ),
            (
                "SOUS-TOTAL Activité : Recrutement et Mobilité",
                "18000000",
                "17400000",
                "20500000",
                "900000",
                "21350000",
                "1200000",
                "0",
                "21700000",
                "",
            ),
            # Activité 2.1.1.2
            ("Activité : Formation et Développement des Compétences", "", "", "", "", "", "", "", "", ""),
            (
                "661400 - Frais de formation continue",
                "2500000",
                "2400000",
                "2900000",
                "125000",
                "3025000",
                "165000",
                "0",
                "3065000",
                "Formations qualifiantes",
            ),
            (
                "661500 - Séminaires et ateliers",
                "1500000",
                "1450000",
                "1800000",
                "75000",
                "1875000",
                "100000",
                "0",
                "1900000",
                "Ateliers de renforcement capacités",
            ),
            (
                "SOUS-TOTAL Activité : Formation et Développement des Compétences",
                "4000000",
                "3850000",
                "4700000",
                "200000",
                "4900000",
                "265000",
                "0",
                "4965000",
                "",
            ),
            (
                "SOUS-TOTAL Service : Direction des Ressources Humaines",
                "22000000",
                "21250000",
                "25200000",
                "1100000",
                "26250000",
                "1465000",
                "0",
                "26665000",
                "",
            ),
            # Service 2.1.2
            ("Service Bénéficiaire : Services Déconcentrés", "", "", "", "", "", "", "", "", ""),
            # Activité 2.1.2.1
            ("Activité : Gestion Personnel Déconcentré", "", "", "", "", "", "", "", "", ""),
            (
                "661100 - Salaires personnel déconcentré",
                "25000000",
                "24000000",
                "28000000",
                "1200000",
                "29200000",
                "1600000",
                "0",
                "29600000",
                "Agents en régions",
            ),
            (
                "661300 - Charges sociales",
                "5000000",
                "4800000",
                "5600000",
                "240000",
                "5840000",
                "320000",
                "0",
                "5920000",
                "CNPS, assurances",
            ),
            (
                "SOUS-TOTAL Activité : Gestion Personnel Déconcentré",
                "30000000",
                "28800000",
                "33600000",
                "1440000",
                "35040000",
                "1920000",
                "0",
                "35520000",
                "",
            ),
            # Activité 2.1.2.2
            ("Activité : Avantages et Œuvres Sociales", "", "", "", "", "", "", "", "", ""),
            (
                "661600 - Allocations familiales",
                "3000000",
                "2900000",
                "3400000",
                "145000",
                "3545000",
                "190000",
                "0",
                "3590000",
                "Allocations",
            ),
            (
                "661700 - Œuvres sociales",
                "2000000",
                "1950000",
                "2300000",
                "95000",
                "2395000",
                "130000",
                "0",
                "2430000",
                "Assistance sociale",
            ),
            (
                "SOUS-TOTAL Activité : Avantages et Œuvres Sociales",
                "5000000",
                "4850000",
                "5700000",
                "240000",
                "5940000",
                "320000",
                "0",
                "6020000",
                "",
            ),
            (
                "SOUS-TOTAL Service : Services Déconcentrés",
                "35000000",
                "33650000",
                "39300000",
                "1680000",
                "40980000",
                "2240000",
                "0",
                "41540000",
                "",
            ),
            (
                "SOUS-TOTAL Action : Gestion des Ressources Humaines",
                "57000000",
                "54900000",
                "64500000",
                "2780000",
                "67230000",
                "3705000",
                "0",
                "68205000",
                "",
            ),
            # Action 2.2
            (
                "Action : Valorisation et Motivation du Personnel",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "Amélioration conditions de travail",
            ),
            # Service 2.2.1
            ("Service Bénéficiaire : Direction du Bien-être au Travail", "", "", "", "", "", "", "", "", ""),
            # Activité 2.2.1.1
            ("Activité : Santé et Sécurité au Travail", "", "", "", "", "", "", "", "", ""),
            (
                "661800 - Assurance maladie du personnel",
                "8000000",
                "7700000",
                "9000000",
                "385000",
                "9385000",
                "515000",
                "0",
                "9515000",
                "Couverture médicale",
            ),
            (
                "661900 - Médecine du travail",
                "1500000",
                "1450000",
                "1700000",
                "75000",
                "1775000",
                "95000",
                "0",
                "1795000",
                "Visites médicales, vaccinations",
            ),
            (
                "SOUS-TOTAL Activité : Santé et Sécurité au Travail",
                "9500000",
                "9150000",
                "10700000",
                "460000",
                "11160000",
                "610000",
                "0",
                "11310000",
                "",
            ),
            # Activité 2.2.1.2
            ("Activité : Motivation et Reconnaissance", "", "", "", "", "", "", "", "", ""),
            (
                "662100 - Primes de performance",
                "5000000",
                "4800000",
                "5600000",
                "240000",
                "5840000",
                "320000",
                "0",
                "5920000",
                "Primes trimestrielles",
            ),
            (
                "662200 - Bonus et gratifications",
                "3000000",
                "2900000",
                "3400000",
                "145000",
                "3545000",
                "190000",
                "0",
                "3590000",
                "Primes de fin d'année",
            ),
            (
                "SOUS-TOTAL Activité : Motivation et Reconnaissance",
                "8000000",
                "7700000",
                "9000000",
                "385000",
                "9385000",
                "510000",
                "0",
                "9510000",
                "",
            ),
            (
                "SOUS-TOTAL Service : Direction du Bien-être au Travail",
                "17500000",
                "16850000",
                "19700000",
                "845000",
                "20545000",
                "1120000",
                "0",
                "20820000",
                "",
            ),
            # Service 2.2.2
            ("Service Bénéficiaire : Direction de la Formation", "", "", "", "", "", "", "", "", ""),
            # Activité 2.2.2.1
            ("Activité : Formation Continue et Perfectionnement", "", "", "", "", "", "", "", "", ""),
            (
                "662400 - Bourses de formation",
                "6000000",
                "5800000",
                "6800000",
                "290000",
                "7090000",
                "385000",
                "0",
                "7185000",
                "Formations diplômantes",
            ),
            (
                "662500 - Stages et séminaires internationaux",
                "4000000",
                "3850000",
                "4500000",
                "195000",
                "4695000",
                "255000",
                "0",
                "4755000",
                "Stages à l'étranger",
            ),
            (
                "SOUS-TOTAL Activité : Formation Continue et Perfectionnement",
                "10000000",
                "9650000",
                "11300000",
                "485000",
                "11785000",
                "640000",
                "0",
                "11940000",
                "",
            ),
            # Activité 2.2.2.2
            ("Activité : Renforcement des Capacités Techniques", "", "", "", "", "", "", "", "", ""),
            (
                "662600 - Formations techniques spécialisées",
                "3500000",
                "3400000",
                "4000000",
                "170000",
                "4170000",
                "225000",
                "0",
                "4225000",
                "Formations métiers",
            ),
            (
                "662700 - Certifications professionnelles",
                "2500000",
                "2400000",
                "2800000",
                "120000",
                "2920000",
                "160000",
                "0",
                "2960000",
                "Certifications ISO, PMP",
            ),
            (
                "SOUS-TOTAL Activité : Renforcement des Capacités Techniques",
                "6000000",
                "5800000",
                "6800000",
                "290000",
                "7090000",
                "385000",
                "0",
                "7185000",
                "",
            ),
            (
                "SOUS-TOTAL Service : Direction de la Formation",
                "16000000",
                "15450000",
                "18100000",
                "775000",
                "18875000",
                "1025000",
                "0",
                "19125000",
                "",
            ),
            (
                "SOUS-TOTAL Action : Valorisation et Motivation du Personnel",
                "33500000",
                "32300000",
                "37800000",
                "1620000",
                "39420000",
                "2145000",
                "0",
                "39945000",
                "",
            ),
            (
                "TOTAL PERSONNEL",
                "90500000",
                "87200000",
                "102300000",
                "4400000",
                "106700000",
                "5850000",
                "0",
                "108150000",
                "",
            ),
            # === NATURE 3 : INVESTISSEMENT ===
            ("", "", "", "", "", "", "", "", "", ""),
            ("INVESTISSEMENT", "", "", "", "", "", "", "", "", "Nature de dépense"),
            # Action 3.1
            (
                "Action : Équipements et Infrastructures",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "Modernisation des équipements",
            ),
            # Service 3.1.1
            ("Service Bénéficiaire : Direction des Infrastructures", "", "", "", "", "", "", "", "", ""),
            # Activité 3.1.1.1
            ("Activité : Construction et Réhabilitation", "", "", "", "", "", "", "", "", ""),
            (
                "221100 - Travaux de construction",
                "50000000",
                "48000000",
                "58000000",
                "2500000",
                "60500000",
                "3300000",
                "0",
                "61300000",
                "Construction nouveaux locaux",
            ),
            (
                "221200 - Réhabilitation bâtiments existants",
                "30000000",
                "29000000",
                "35000000",
                "1500000",
                "36500000",
                "2000000",
                "0",
                "37000000",
                "Rénovation bureaux",
            ),
            (
                "SOUS-TOTAL Activité : Construction et Réhabilitation",
                "80000000",
                "77000000",
                "93000000",
                "4000000",
                "97000000",
                "5300000",
                "0",
                "98300000",
                "",
            ),
            # Activité 3.1.1.2
            ("Activité : Aménagements et Installations", "", "", "", "", "", "", "", "", ""),
            (
                "221300 - Aménagements extérieurs",
                "15000000",
                "14500000",
                "17500000",
                "750000",
                "18250000",
                "1000000",
                "0",
                "18500000",
                "Parkings, espaces verts",
            ),
            (
                "221400 - Installations techniques",
                "10000000",
                "9700000",
                "11500000",
                "490000",
                "11990000",
                "650000",
                "0",
                "12150000",
                "Électricité, plomberie",
            ),
            (
                "SOUS-TOTAL Activité : Aménagements et Installations",
                "25000000",
                "24200000",
                "29000000",
                "1240000",
                "30240000",
                "1650000",
                "0",
                "30650000",
                "",
            ),
            (
                "SOUS-TOTAL Service : Direction des Infrastructures",
                "105000000",
                "101200000",
                "122000000",
                "5240000",
                "127240000",
                "6950000",
                "0",
                "128950000",
                "",
            ),
            # Service 3.1.2
            ("Service Bénéficiaire : Direction des Équipements", "", "", "", "", "", "", "", "", ""),
            # Activité 3.1.2.1
            ("Activité : Acquisition Matériels et Équipements", "", "", "", "", "", "", "", "", ""),
            (
                "244100 - Mobilier de bureau",
                "12000000",
                "11600000",
                "14000000",
                "600000",
                "14600000",
                "800000",
                "0",
                "14800000",
                "Bureaux, chaises, armoires",
            ),
            (
                "244200 - Matériel informatique",
                "20000000",
                "19300000",
                "23000000",
                "1000000",
                "24000000",
                "1300000",
                "0",
                "24300000",
                "PC, imprimantes, serveurs",
            ),
            (
                "SOUS-TOTAL Activité : Acquisition Matériels et Équipements",
                "32000000",
                "30900000",
                "37000000",
                "1600000",
                "38600000",
                "2100000",
                "0",
                "39100000",
                "",
            ),
            # Activité 3.1.2.2
            ("Activité : Véhicules et Engins", "", "", "", "", "", "", "", "", ""),
            (
                "245100 - Acquisition véhicules de service",
                "35000000",
                "33800000",
                "40000000",
                "1720000",
                "41720000",
                "2280000",
                "0",
                "42280000",
                "Véhicules 4x4, berlines",
            ),
            (
                "245200 - Engins et matériel roulant",
                "18000000",
                "17400000",
                "21000000",
                "900000",
                "21900000",
                "1200000",
                "0",
                "22200000",
                "Engins de chantier",
            ),
            (
                "SOUS-TOTAL Activité : Véhicules et Engins",
                "53000000",
                "51200000",
                "61000000",
                "2620000",
                "63620000",
                "3480000",
                "0",
                "64480000",
                "",
            ),
            (
                "SOUS-TOTAL Service : Direction des Équipements",
                "85000000",
                "82100000",
                "98000000",
                "4220000",
                "102220000",
                "5580000",
                "0",
                "103580000",
                "",
            ),
            (
                "SOUS-TOTAL Action : Équipements et Infrastructures",
                "190000000",
                "183300000",
                "220000000",
                "9460000",
                "229460000",
                "12530000",
                "0",
                "232530000",
                "",
            ),
            # Action 3.2
            ("Action : Modernisation et Digitalisation", "", "", "", "", "", "", "", "", "Transformation numérique"),
            # Service 3.2.1
            ("Service Bénéficiaire : Direction de la Transformation Numérique", "", "", "", "", "", "", "", "", ""),
            # Activité 3.2.1.1
            ("Activité : Systèmes d'Information", "", "", "", "", "", "", "", "", ""),
            (
                "218300 - Logiciels de gestion intégrés (ERP)",
                "25000000",
                "24000000",
                "29000000",
                "1250000",
                "30250000",
                "1650000",
                "0",
                "30650000",
                "SAP, Oracle, Microsoft Dynamics",
            ),
            (
                "218400 - Infrastructure cloud et cybersécurité",
                "15000000",
                "14500000",
                "17500000",
                "750000",
                "18250000",
                "1000000",
                "0",
                "18500000",
                "Serveurs cloud, pare-feu",
            ),
            (
                "SOUS-TOTAL Activité : Systèmes d'Information",
                "40000000",
                "38500000",
                "46500000",
                "2000000",
                "48500000",
                "2650000",
                "0",
                "49150000",
                "",
            ),
            # Activité 3.2.1.2
            ("Activité : Équipements Technologiques", "", "", "", "", "", "", "", "", ""),
            (
                "218500 - Équipements audiovisuels",
                "8000000",
                "7700000",
                "9200000",
                "395000",
                "9595000",
                "525000",
                "0",
                "9725000",
                "Vidéoprojecteurs, écrans",
            ),
            (
                "218600 - Solutions de visioconférence",
                "6000000",
                "5800000",
                "7000000",
                "300000",
                "7300000",
                "400000",
                "0",
                "7400000",
                "Systèmes de réunion à distance",
            ),
            (
                "SOUS-TOTAL Activité : Équipements Technologiques",
                "14000000",
                "13500000",
                "16200000",
                "695000",
                "16895000",
                "925000",
                "0",
                "17125000",
                "",
            ),
            (
                "SOUS-TOTAL Service : Direction de la Transformation Numérique",
                "54000000",
                "52000000",
                "62700000",
                "2695000",
                "65395000",
                "3575000",
                "0",
                "66275000",
                "",
            ),
            # Service 3.2.2
            ("Service Bénéficiaire : Direction de l'Innovation", "", "", "", "", "", "", "", "", ""),
            # Activité 3.2.2.1
            ("Activité : Recherche et Développement", "", "", "", "", "", "", "", "", ""),
            (
                "218700 - Équipements de laboratoire",
                "12000000",
                "11600000",
                "14000000",
                "600000",
                "14600000",
                "800000",
                "0",
                "14800000",
                "Matériel scientifique",
            ),
            (
                "218800 - Prototypes et pilots",
                "8000000",
                "7700000",
                "9200000",
                "395000",
                "9595000",
                "525000",
                "0",
                "9725000",
                "Projets pilotes",
            ),
            (
                "SOUS-TOTAL Activité : Recherche et Développement",
                "20000000",
                "19300000",
                "23200000",
                "995000",
                "24195000",
                "1325000",
                "0",
                "24525000",
                "",
            ),
            # Activité 3.2.2.2
            ("Activité : Veille Technologique et Innovation", "", "", "", "", "", "", "", "", ""),
            (
                "218900 - Abonnements bases de données techniques",
                "5000000",
                "4800000",
                "5800000",
                "250000",
                "6050000",
                "330000",
                "0",
                "6130000",
                "Bases de données scientifiques",
            ),
            (
                "219000 - Partenariats innovation",
                "7000000",
                "6750000",
                "8100000",
                "350000",
                "8450000",
                "460000",
                "0",
                "8560000",
                "Collaborations universités",
            ),
            (
                "SOUS-TOTAL Activité : Veille Technologique et Innovation",
                "12000000",
                "11550000",
                "13900000",
                "600000",
                "14500000",
                "790000",
                "0",
                "14690000",
                "",
            ),
            (
                "SOUS-TOTAL Service : Direction de l'Innovation",
                "32000000",
                "30850000",
                "37100000",
                "1595000",
                "38695000",
                "2115000",
                "0",
                "39215000",
                "",
            ),
            (
                "SOUS-TOTAL Action : Modernisation et Digitalisation",
                "86000000",
                "82850000",
                "99800000",
                "4290000",
                "104090000",
                "5690000",
                "0",
                "105490000",
                "",
            ),
            (
                "TOTAL INVESTISSEMENT",
                "276000000",
                "266150000",
                "319800000",
                "13750000",
                "333550000",
                "18220000",
                "0",
                "338020000",
                "",
            ),
            # === TOTAL GÉNÉRAL ===
            ("", "", "", "", "", "", "", "", "", ""),
            (
                "TOTAL GÉNÉRAL",
                "387400000",
                "373290000",
                "445800000",
                "19185000",
                "464985000",
                "25415000",
                "0",
                "471215000",
                "",
            ),
        ]

        current_row = 2

        # Définir les couleurs selon la hiérarchie (comme dans le PDF)
        nature_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Or (jaune doré)
        action_fill = PatternFill(start_color="9BC2E6", end_color="9BC2E6", fill_type="solid")  # Bleu clair
        service_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")  # Jaune orangé
        activite_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")  # Vert clair
        ligne_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")  # Blanc
        sous_total_activite_fill = PatternFill(
            start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"
        )  # Vert très clair
        sous_total_service_fill = PatternFill(
            start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
        )  # Jaune très clair
        sous_total_action_fill = PatternFill(
            start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"
        )  # Bleu très clair
        total_nature_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Or
        total_general_fill = PatternFill(start_color="DC3545", end_color="DC3545", fill_type="solid")  # Rouge

        for idx, exemple in enumerate(exemples):
            libelle = exemple[0]

            # Déterminer le type de ligne et appliquer la couleur
            if libelle.upper() in ["BIENS ET SERVICES", "PERSONNEL", "INVESTISSEMENT", "INVESTISSEMENTS", "TRANSFERTS"]:
                row_fill = nature_fill
                row_font = Font(bold=True, size=11)
            elif libelle.startswith("Action :") or libelle.startswith("- Action :"):
                row_fill = action_fill
                row_font = Font(bold=True, size=10)
            elif libelle.startswith("Service Bénéficiaire :") or libelle.startswith("- Service Bénéficiaire :"):
                row_fill = service_fill
                row_font = Font(bold=True, size=9)
            elif libelle.startswith("Activité :") or libelle.startswith("- Activité :"):
                row_fill = activite_fill
                row_font = Font(bold=True, size=9)
            elif libelle.startswith("SOUS-TOTAL Activité"):
                row_fill = sous_total_activite_fill
                row_font = Font(bold=True, size=8, italic=True)
            elif libelle.startswith("SOUS-TOTAL Service"):
                row_fill = sous_total_service_fill
                row_font = Font(bold=True, size=9, italic=True)
            elif libelle.startswith("SOUS-TOTAL Action"):
                row_fill = sous_total_action_fill
                row_font = Font(bold=True, size=10, italic=True)
            elif libelle.startswith("TOTAL ") and not libelle.startswith("TOTAL GÉNÉRAL"):
                row_fill = total_nature_fill
                row_font = Font(bold=True, size=11)
            elif libelle.startswith("TOTAL GÉNÉRAL"):
                row_fill = total_general_fill
                row_font = Font(bold=True, size=12, color="FFFFFF")
            elif libelle == "":
                row_fill = PatternFill()  # Transparent
                row_font = Font(size=9)
            else:
                # Ligne budgétaire normale
                row_fill = ligne_fill
                row_font = Font(size=9)

            for col_num, value in enumerate(exemple, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border_style
                cell.fill = row_fill
                cell.font = row_font

                # Aligner les nombres à droite
                if col_num > 1 and col_num < 10 and value and value != "":
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

                # Formater les nombres
                if col_num > 1 and col_num < 10 and value and str(value).replace(" ", "").isdigit():
                    cell.number_format = "#,##0"

            current_row += 1

        # Ajouter une feuille d'instructions
        ws_instructions = wb.create_sheet("📋 Instructions")
        ws_instructions.column_dimensions["A"].width = 80

        instructions = [
            ("MODÈLE DE FICHE TECHNIQUE BUDGÉTAIRE", header_font, header_fill),
            ("", None, None),
            ("📅 NOTATION TEMPORELLE", Font(bold=True, size=12, color="2196F3"), None),
            ("", None, None),
            ("N = Année en cours (année de référence)", None, None),
            ("N+1 = Année budgétaire à préparer (année suivante)", None, None),
            ("", None, None),
            (
                "Exemple : Si vous préparez le budget 2025, alors N = 2024 et N+1 = 2025",
                Font(italic=True),
                example_fill,
            ),
            ("", None, None),
            ("📌 STRUCTURE DU FICHIER", Font(bold=True, size=12), None),
            ("", None, None),
            ("Ce fichier doit respecter une hiérarchie stricte :", None, None),
            ("", None, None),
            ("1️⃣ NATURE DE DÉPENSE : BIENS ET SERVICES, PERSONNEL, INVESTISSEMENT, ou TRANSFERTS", None, example_fill),
            ("   ↓", None, None),
            ("2️⃣ ACTION : Commencer par 'Action :' ou '- Action :'", None, example_fill),
            ("   ↓", None, None),
            (
                "3️⃣ SERVICE BÉNÉFICIAIRE : Commencer par 'Service Bénéficiaire :' ou '- Service Bénéficiaire :'",
                None,
                example_fill,
            ),
            ("   ↓", None, None),
            ("4️⃣ ACTIVITÉ : Commencer par 'Activité :' ou '- Activité :'", None, example_fill),
            ("   ↓", None, None),
            ("5️⃣ LIGNES BUDGÉTAIRES : Détails des dépenses (sans préfixe spécial)", None, example_fill),
            ("", None, None),
            ("", None, None),
            ("⚠️ RÈGLES IMPORTANTES", Font(bold=True, size=12, color="DC3545"), None),
            ("", None, None),
            ("✅ Les montants doivent être des nombres (sans espace ni symbole)", None, None),
            ("✅ La colonne 'CODE / LIBELLE' ne doit JAMAIS être vide", None, None),
            ("✅ Respectez exactement les préfixes : 'Action :', 'Service Bénéficiaire :', 'Activité :'", None, None),
            ("✅ Les natures de dépenses doivent être en MAJUSCULES", None, None),
            ("✅ Ne supprimez pas les en-têtes de colonnes", None, None),
            ("", None, None),
            ("", None, None),
            ("💡 EXEMPLES", Font(bold=True, size=12, color="28A745"), None),
            ("", None, None),
            ("Consultez la feuille 'Fiche Technique' pour voir des exemples concrets.", None, None),
            ("Les lignes en gris sont des exemples à SUPPRIMER avant de charger votre fichier.", None, None),
            ("", None, None),
            ("", None, None),
            (f"📅 Modèle généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", Font(italic=True, size=9), None),
        ]

        for row_num, (text, font, fill) in enumerate(instructions, 1):
            cell = ws_instructions.cell(row=row_num, column=1)
            cell.value = text
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Figer la première ligne (en-têtes uniquement) dans la feuille Fiche Technique
        ws.freeze_panes = "A2"

        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Log de l'activité
        logger.info(f"📥 Template de fiche technique téléchargé par {current_user.email}")

        # Retourner le fichier (nom générique sans année)
        filename = "Modele_Fiche_Technique_N.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        logger.error(f"Erreur génération template: {e}")
        raise HTTPException(500, f"Erreur lors de la génération du modèle: {e!s}")


@router.post("/api/charger-fiche")
async def api_charger_fiche(
    fichier: UploadFile = File(...),
    programme_id: int = Form(...),
    annee: int = Form(...),
    nom_fiche: str | None = Form(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """
    Charger et analyser une fiche technique depuis un fichier Excel/PDF
    """
    try:
        # Vérifier que le programme existe
        programme = session.get(Programme, programme_id)
        if not programme:
            raise HTTPException(400, "Programme non trouvé")

        # Lire le fichier
        content = await fichier.read()

        # Déterminer le type de fichier
        if fichier.filename.endswith(".pdf"):
            return await _analyser_fiche_pdf(content, nom_fiche, programme_id, annee, session, current_user)
        elif fichier.filename.endswith((".xlsx", ".xls")):
            return await _analyser_fiche_excel(content, nom_fiche, programme_id, annee, session, current_user)
        else:
            raise HTTPException(400, "Format de fichier non supporté. Utilisez Excel (.xlsx, .xls) ou PDF (.pdf)")

    except Exception as e:
        logger.error(f"Erreur chargement fiche: {e}")
        raise HTTPException(500, f"Erreur lors du chargement: {e!s}")


async def _analyser_fiche_excel(
    content: bytes, nom_fiche: str | None, programme_id: int, annee: int, session: Session, current_user: User
):
    """
    Analyser un fichier Excel de fiche technique avec le template standardisé
    Utilise le FicheTechniqueService pour toute la logique métier
    """
    return FicheTechniqueService.analyser_fichier_excel(
        content=content,
        nom_fiche=nom_fiche,
        programme_id=programme_id,
        annee=annee,
        session=session,
        current_user=current_user,
    )


async def _analyser_fiche_pdf(
    content: bytes, nom_fiche: str | None, programme_id: int, annee: int, session: Session, current_user: User
):
    """
    Analyser un fichier PDF de fiche technique et extraire la structure hiérarchique
    """
    try:
        from io import BytesIO

        import PyPDF2
    except ImportError:
        raise HTTPException(
            500,
            "❌ PyPDF2 n'est pas installé.\n\n"
            "💡 Installez-le avec : uv add pypdf2\n\n"
            "📥 Ou utilisez le modèle Excel pour plus de fiabilité.",
        )

    try:
        # Extraire le texte du PDF
        pdf_reader = PyPDF2.PdfReader(BytesIO(content))

        logger.info(f"📄 PDF chargé : {len(pdf_reader.pages)} page(s)")

        # Extraire tout le texte
        full_text = ""
        for page_num, page in enumerate(pdf_reader.pages):
            text = page.extract_text()
            full_text += text + "\n"
            logger.debug(f"  📄 Page {page_num + 1} : {len(text)} caractères")

        if not full_text.strip():
            raise HTTPException(
                400, "❌ Impossible d'extraire le texte du PDF. Le fichier est peut-être scanné ou protégé."
            )

        logger.info(f"✅ Texte extrait : {len(full_text)} caractères")

        # Créer un DataFrame à partir du texte parsé
        df_data = []
        lignes_texte = full_text.split("\n")

        current_nature = None
        current_action = None
        current_service = None
        current_activite = None

        for ligne_idx, ligne in enumerate(lignes_texte):
            ligne = ligne.strip()
            if not ligne:
                continue

            # Détecter les natures de dépenses
            if ligne.upper() in ["BIENS ET SERVICES", "PERSONNEL", "INVESTISSEMENT", "INVESTISSEMENTS", "TRANSFERTS"]:
                current_nature = ligne.upper()
                df_data.append({"type": "nature", "nature": current_nature, "libelle": ligne, "montants": {}})
                logger.debug(f"📌 Nature : {current_nature}")
                continue

            # Détecter les actions (avec pattern flexible)
            if any(ligne.startswith(p) for p in ["Action :", "- Action :", "ACTION :"]):
                action_libelle = ligne.replace("Action :", "").replace("- Action :", "").replace("ACTION :", "").strip()
                current_action = action_libelle
                df_data.append(
                    {
                        "type": "action",
                        "nature": current_nature,
                        "libelle": action_libelle,
                        "montants": _extraire_montants_ligne(ligne),
                    }
                )
                logger.debug(f"  → Action : {action_libelle[:50]}")
                continue

            # Détecter les services
            if any(ligne.startswith(p) for p in ["Service Bénéficiaire :", "- Service Bénéficiaire :", "SERVICE :"]):
                service_libelle = (
                    ligne.replace("Service Bénéficiaire :", "")
                    .replace("- Service Bénéficiaire :", "")
                    .replace("SERVICE :", "")
                    .strip()
                )
                current_service = service_libelle
                df_data.append(
                    {
                        "type": "service",
                        "nature": current_nature,
                        "action": current_action,
                        "libelle": service_libelle,
                        "montants": {},
                    }
                )
                logger.debug(f"    → Service : {service_libelle[:50]}")
                continue

            # Détecter les activités
            if any(ligne.startswith(p) for p in ["Activité :", "- Activité :", "ACTIVITÉ :", "ACTIVITE :"]):
                activite_libelle = (
                    ligne.replace("Activité :", "")
                    .replace("- Activité :", "")
                    .replace("ACTIVITÉ :", "")
                    .replace("ACTIVITE :", "")
                    .strip()
                )
                current_activite = activite_libelle
                df_data.append(
                    {
                        "type": "activite",
                        "nature": current_nature,
                        "action": current_action,
                        "service": current_service,
                        "libelle": activite_libelle,
                        "montants": _extraire_montants_ligne(ligne),
                    }
                )
                logger.debug(f"      → Activité : {activite_libelle[:50]}")
                continue

            # Détecter les lignes budgétaires (commence par un numéro de compte)
            if ligne and ligne[0].isdigit() and len(ligne) > 5:
                df_data.append(
                    {
                        "type": "ligne",
                        "nature": current_nature,
                        "action": current_action,
                        "service": current_service,
                        "activite": current_activite,
                        "libelle": ligne,
                        "montants": _extraire_montants_ligne(ligne),
                    }
                )
                logger.debug(f"        → Ligne : {ligne[:50]}")

        logger.info(f"✅ {len(df_data)} éléments extraits du PDF")

        # Créer la fiche et la structure
        prog = session.get(Programme, programme_id)
        if not prog:
            raise HTTPException(400, "Programme non trouvé")

        # Générer numéro de fiche
        count = session.exec(select(func.count(FicheTechnique.id)).where(FicheTechnique.annee_budget == annee)).one()

        numero_fiche = nom_fiche or f"FT-{annee}-{prog.code}-{count + 1:03d}"

        fiche = FicheTechnique(
            numero_fiche=numero_fiche,
            annee_budget=annee,
            programme_id=programme_id,
            direction_id=None,
            budget_total_demande=Decimal("0"),
            statut="Brouillon",
            phase="Conférence interne",
            created_by_user_id=current_user.id,
        )

        session.add(fiche)
        session.commit()
        session.refresh(fiche)

        logger.info(f"✅ Fiche créée : {fiche.numero_fiche}")

        # Créer la structure depuis les données extraites
        result = _creer_structure_depuis_pdf_data(df_data, fiche.id, session)

        # Recalculer les totaux
        FicheTechniqueService._recalculer_totaux_hierarchie(fiche.id, session)

        # Mettre à jour le budget total
        fiche_updated = session.get(FicheTechnique, fiche.id)
        if fiche_updated:
            budget_total = sum(
                action.projet_budget_n_plus_1 or Decimal("0")
                for action in session.exec(
                    select(ActionBudgetaire).where(ActionBudgetaire.fiche_technique_id == fiche.id)
                ).all()
            )
            fiche_updated.budget_total_demande = budget_total
            session.add(fiche_updated)
            session.commit()

        return {
            "ok": True,
            "fiche_numero": fiche.numero_fiche,
            "actions_count": result["actions_count"],
            "services_count": result["services_count"],
            "activites_count": result["activites_count"],
            "lignes_count": result["lignes_count"],
            "budget_total": float(budget_total),
            "errors": result.get("errors", []),
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erreur analyse PDF : {e}")
        raise HTTPException(500, f"Erreur lors de l'analyse du PDF : {e!s}")


def _extraire_montants_ligne(ligne: str) -> dict:
    """
    Extraire les montants d'une ligne de texte PDF
    Recherche des nombres (avec ou sans séparateurs)
    """
    import re

    # Pattern pour trouver les nombres (avec espaces, virgules, points)
    pattern = r"[\d\s,\.]+(?=\s|$)"
    montants_trouves = re.findall(pattern, ligne)

    montants = {}
    colonnes = [
        "budget_vote_n",
        "budget_actuel_n",
        "enveloppe_n_plus_1",
        "complement_solicite",
        "budget_souhaite",
        "engagement_etat",
        "autre_complement",
        "projet_budget_n_plus_1",
    ]

    for idx, montant_str in enumerate(montants_trouves):
        if idx < len(colonnes):
            try:
                # Nettoyer le montant
                montant_clean = montant_str.replace(" ", "").replace(",", "").replace(".", "")
                if montant_clean.isdigit():
                    montants[colonnes[idx]] = Decimal(montant_clean)
            except:
                pass

    return montants


def _creer_structure_depuis_pdf_data(df_data: list, fiche_id: int, session: Session) -> dict:
    """
    Créer la structure hiérarchique depuis les données extraites du PDF
    """
    actions_count = 0
    services_count = 0
    activites_count = 0
    lignes_count = 0
    errors = []

    actions_map = {}  # {libelle_action: action_obj}
    services_map = {}  # {(libelle_action, libelle_service): service_obj}
    activites_map = {}  # {(libelle_action, libelle_service, libelle_activite): activite_obj}

    for idx, item in enumerate(df_data):
        try:
            if item["type"] == "nature":
                # Juste pour le contexte, pas d'objet à créer
                continue

            elif item["type"] == "action":
                action_key = item["libelle"]
                if action_key not in actions_map:
                    # Générer code
                    code = f"ACT_{actions_count + 1:03d}"

                    action = ActionBudgetaire(
                        fiche_technique_id=fiche_id,
                        nature_depense=item.get("nature"),
                        code=code,
                        libelle=item["libelle"],
                        budget_vote_n=item["montants"].get("budget_vote_n", Decimal("0")),
                        budget_actuel_n=item["montants"].get("budget_actuel_n", Decimal("0")),
                        enveloppe_n_plus_1=item["montants"].get("enveloppe_n_plus_1", Decimal("0")),
                        complement_solicite=item["montants"].get("complement_solicite", Decimal("0")),
                        budget_souhaite=item["montants"].get("budget_souhaite", Decimal("0")),
                        engagement_etat=item["montants"].get("engagement_etat", Decimal("0")),
                        autre_complement=item["montants"].get("autre_complement", Decimal("0")),
                        projet_budget_n_plus_1=item["montants"].get("projet_budget_n_plus_1", Decimal("0")),
                        ordre=actions_count,
                    )
                    session.add(action)
                    session.commit()
                    session.refresh(action)
                    actions_map[action_key] = action
                    actions_count += 1
                    logger.debug(f"  ✅ Action créée : {item['libelle'][:50]}")

            elif item["type"] == "service":
                service_key = (item.get("action"), item["libelle"])
                if service_key not in services_map:
                    action_parent = actions_map.get(item.get("action"))
                    if action_parent:
                        code = f"SRV_{services_count + 1:03d}"

                        service = ServiceBeneficiaire(
                            fiche_technique_id=fiche_id,
                            action_id=action_parent.id,
                            code=code,
                            libelle=item["libelle"],
                            ordre=services_count,
                        )
                        session.add(service)
                        session.commit()
                        session.refresh(service)
                        services_map[service_key] = service
                        services_count += 1
                        logger.debug(f"    ✅ Service créé : {item['libelle'][:50]}")
                    else:
                        errors.append(f"Service sans action : {item['libelle']}")

            elif item["type"] == "activite":
                activite_key = (item.get("action"), item.get("service"), item["libelle"])
                if activite_key not in activites_map:
                    service_parent = services_map.get((item.get("action"), item.get("service")))
                    if service_parent:
                        code = f"ACTIV_{activites_count + 1:03d}"

                        activite = ActiviteBudgetaire(
                            fiche_technique_id=fiche_id,
                            service_beneficiaire_id=service_parent.id,
                            code=code,
                            libelle=item["libelle"],
                            budget_vote_n=item["montants"].get("budget_vote_n", Decimal("0")),
                            budget_actuel_n=item["montants"].get("budget_actuel_n", Decimal("0")),
                            enveloppe_n_plus_1=item["montants"].get("enveloppe_n_plus_1", Decimal("0")),
                            complement_solicite=item["montants"].get("complement_solicite", Decimal("0")),
                            budget_souhaite=item["montants"].get("budget_souhaite", Decimal("0")),
                            engagement_etat=item["montants"].get("engagement_etat", Decimal("0")),
                            autre_complement=item["montants"].get("autre_complement", Decimal("0")),
                            projet_budget_n_plus_1=item["montants"].get("projet_budget_n_plus_1", Decimal("0")),
                            ordre=activites_count,
                        )
                        session.add(activite)
                        session.commit()
                        session.refresh(activite)
                        activites_map[activite_key] = activite
                        activites_count += 1
                        logger.debug(f"      ✅ Activité créée : {item['libelle'][:50]}")
                    else:
                        errors.append(f"Activité sans service : {item['libelle']}")

            elif item["type"] == "ligne":
                activite_parent = activites_map.get((item.get("action"), item.get("service"), item.get("activite")))
                if activite_parent:
                    # Extraire le code du début de la ligne
                    libelle = item["libelle"]
                    code_match = re.match(r"^(\d+)", libelle)
                    code = code_match.group(1) if code_match else f"LIGNE_{lignes_count + 1:05d}"

                    ligne = LigneBudgetaireDetail(
                        fiche_technique_id=fiche_id,
                        activite_id=activite_parent.id,
                        code=code,
                        libelle=libelle,
                        budget_vote_n=item["montants"].get("budget_vote_n", Decimal("0")),
                        budget_actuel_n=item["montants"].get("budget_actuel_n", Decimal("0")),
                        enveloppe_n_plus_1=item["montants"].get("enveloppe_n_plus_1", Decimal("0")),
                        complement_solicite=item["montants"].get("complement_solicite", Decimal("0")),
                        budget_souhaite=item["montants"].get("budget_souhaite", Decimal("0")),
                        engagement_etat=item["montants"].get("engagement_etat", Decimal("0")),
                        autre_complement=item["montants"].get("autre_complement", Decimal("0")),
                        projet_budget_n_plus_1=item["montants"].get("projet_budget_n_plus_1", Decimal("0")),
                        ordre=lignes_count,
                    )
                    session.add(ligne)
                    session.commit()
                    session.refresh(ligne)
                    lignes_count += 1
                    logger.debug(f"        ✅ Ligne créée : {libelle[:50]}")
                else:
                    errors.append(f"Ligne sans activité : {item['libelle']}")

        except Exception as e:
            errors.append(f"Erreur élément {idx}: {e!s}")
            logger.error(f"❌ Erreur élément {idx}: {e}")

    logger.info(
        f"📊 Structure créée : {actions_count} actions, {services_count} services, {activites_count} activités, {lignes_count} lignes"
    )

    return {
        "actions_count": actions_count,
        "services_count": services_count,
        "activites_count": activites_count,
        "lignes_count": lignes_count,
        "errors": errors,
    }


# ============================================
# CRÉATION DES ÉLÉMENTS HIÉRARCHIQUES
# ============================================


@router.post("/api/actions")
async def api_create_action(
    fiche_id: int = Form(...),
    nature_depense: str = Form(...),
    code: str = Form(...),
    libelle: str = Form(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Créer une nouvelle action"""
    fiche = session.get(FicheTechnique, fiche_id)
    if not fiche:
        raise HTTPException(404, "Fiche technique non trouvée")

    try:
        # Déterminer l'ordre
        max_ordre = (
            session.exec(
                select(func.max(ActionBudgetaire.ordre)).where(ActionBudgetaire.fiche_technique_id == fiche.id)
            ).one()
            or 0
        )

        action = ActionBudgetaire(
            fiche_technique_id=fiche.id,
            nature_depense=nature_depense,
            code=code,
            libelle=libelle,
            ordre=max_ordre + 1,
            budget_vote_n=Decimal("0"),
            budget_actuel_n=Decimal("0"),
            enveloppe_n_plus_1=Decimal("0"),
            complement_solicite=Decimal("0"),
            budget_souhaite=Decimal("0"),
            engagement_etat=Decimal("0"),
            autre_complement=Decimal("0"),
            projet_budget_n_plus_1=Decimal("0"),
        )

        session.add(action)
        session.commit()
        session.refresh(action)

        logger.info(f"✅ Action {code} créée par {current_user.email}")
        return {"ok": True, "id": action.id, "message": "Action créée avec succès"}
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur création action: {e}")
        raise HTTPException(500, f"Erreur lors de la création: {e!s}")


@router.post("/api/services")
async def api_create_service(
    action_id: int = Form(...),
    code: str = Form(...),
    libelle: str = Form(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Créer un nouveau service bénéficiaire"""
    action = session.get(ActionBudgetaire, action_id)
    if not action:
        raise HTTPException(404, "Action parente non trouvée")

    try:
        # Déterminer l'ordre
        max_ordre = (
            session.exec(
                select(func.max(ServiceBeneficiaire.ordre)).where(ServiceBeneficiaire.action_id == action_id)
            ).one()
            or 0
        )

        service = ServiceBeneficiaire(
            fiche_technique_id=action.fiche_technique_id,
            action_id=action_id,
            code=code,
            libelle=libelle,
            ordre=max_ordre + 1,
        )

        session.add(service)
        session.commit()
        session.refresh(service)

        logger.info(f"✅ Service {code} créé par {current_user.email}")
        return {"ok": True, "id": service.id, "message": "Service créé avec succès"}
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur création service: {e}")
        raise HTTPException(500, f"Erreur lors de la création: {e!s}")


@router.post("/api/activites")
async def api_create_activite(
    service_id: int = Form(...),
    code: str = Form(...),
    libelle: str = Form(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Créer une nouvelle activité"""
    service = session.get(ServiceBeneficiaire, service_id)
    if not service:
        raise HTTPException(404, "Service parent non trouvé")

    try:
        # Déterminer l'ordre
        max_ordre = (
            session.exec(
                select(func.max(ActiviteBudgetaire.ordre)).where(
                    ActiviteBudgetaire.service_beneficiaire_id == service_id
                )
            ).one()
            or 0
        )

        activite = ActiviteBudgetaire(
            fiche_technique_id=service.fiche_technique_id,
            service_beneficiaire_id=service_id,
            code=code,
            libelle=libelle,
            ordre=max_ordre + 1,
            budget_vote_n=Decimal("0"),
            budget_actuel_n=Decimal("0"),
            enveloppe_n_plus_1=Decimal("0"),
            complement_solicite=Decimal("0"),
            budget_souhaite=Decimal("0"),
            engagement_etat=Decimal("0"),
            autre_complement=Decimal("0"),
            projet_budget_n_plus_1=Decimal("0"),
        )

        session.add(activite)
        session.commit()
        session.refresh(activite)

        logger.info(f"✅ Activité {code} créée par {current_user.email}")
        return {"ok": True, "id": activite.id, "message": "Activité créée avec succès"}
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur création activité: {e}")
        raise HTTPException(500, f"Erreur lors de la création: {e!s}")


@router.post("/api/lignes")
async def api_create_ligne(
    activite_id: int = Form(...),
    code: str = Form(...),
    libelle: str = Form(...),
    budget_vote_n: float = Form(0),
    budget_actuel_n: float = Form(0),
    enveloppe_n_plus_1: float = Form(0),
    complement_solicite: float = Form(0),
    engagement_etat: float = Form(0),
    autre_complement: float = Form(0),
    justificatifs: str | None = Form(None),
    documents: list[UploadFile] = File(default=[]),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Créer une nouvelle ligne budgétaire avec documents"""
    activite = session.get(ActiviteBudgetaire, activite_id)
    if not activite:
        raise HTTPException(404, "Activité parente non trouvée")

    try:
        # Récupérer les codes pour le renommage des fichiers
        service = session.get(ServiceBeneficiaire, activite.service_beneficiaire_id)
        action = session.get(ActionBudgetaire, service.action_id)

        # Déterminer l'ordre
        max_ordre = (
            session.exec(
                select(func.max(LigneBudgetaireDetail.ordre)).where(LigneBudgetaireDetail.activite_id == activite_id)
            ).one()
            or 0
        )

        # Créer la ligne budgétaire
        ligne = LigneBudgetaireDetail(
            fiche_technique_id=activite.fiche_technique_id,
            activite_id=activite_id,
            code=code,
            libelle=libelle,
            ordre=max_ordre + 1,
            budget_vote_n=Decimal(str(budget_vote_n)),
            budget_actuel_n=Decimal(str(budget_actuel_n)),
            enveloppe_n_plus_1=Decimal(str(enveloppe_n_plus_1)),
            complement_solicite=Decimal(str(complement_solicite)),
            engagement_etat=Decimal(str(engagement_etat)),
            autre_complement=Decimal(str(autre_complement)),
            justificatifs=justificatifs,
        )

        # Calculer les champs dérivés
        ligne.budget_souhaite = ligne.enveloppe_n_plus_1 + ligne.complement_solicite
        ligne.projet_budget_n_plus_1 = ligne.enveloppe_n_plus_1 + ligne.engagement_etat + ligne.autre_complement

        session.add(ligne)
        session.commit()
        session.refresh(ligne)

        # Gérer l'upload des documents
        documents_count = 0
        if documents and len(documents) > 0 and documents[0].filename:
            docs_dir = Path(f"uploads/budget/lignes/{ligne.id}")
            docs_dir.mkdir(parents=True, exist_ok=True)

            for doc in documents:
                if doc.filename:
                    # Renommer le fichier selon le format : CodeAction_CodeActivité_CodeLigne_NomOriginal.ext
                    file_ext = Path(doc.filename).suffix
                    original_name = Path(doc.filename).stem
                    new_filename = f"{action.code}_{activite.code}_{code}_{original_name}{file_ext}"

                    # Sauvegarder le fichier
                    file_path = docs_dir / new_filename
                    content = await doc.read()

                    with open(file_path, "wb") as f:
                        f.write(content)

                    # Enregistrer les métadonnées en base
                    doc_meta = DocumentLigneBudgetaire(
                        ligne_budgetaire_id=ligne.id,
                        fiche_technique_id=activite.fiche_technique_id,
                        nom_fichier_original=doc.filename,
                        nom_fichier_stocke=new_filename,
                        chemin_fichier=str(file_path),
                        type_fichier=file_ext,
                        taille_octets=len(content),
                        code_action=action.code,
                        code_activite=activite.code,
                        code_ligne=code,
                        uploaded_by_user_id=current_user.id,
                    )
                    session.add(doc_meta)

                    documents_count += 1
                    logger.info(f"📎 Document sauvegardé : {new_filename}")

            # Committer les métadonnées des documents
            session.commit()

        # Recalculer les totaux de toute la hiérarchie
        FicheTechniqueService._recalculer_totaux_hierarchie(activite.fiche_technique_id, session)

        logger.info(f"✅ Ligne budgétaire {code} créée avec {documents_count} document(s) par {current_user.email}")
        return {
            "ok": True,
            "id": ligne.id,
            "documents_count": documents_count,
            "message": f"Ligne budgétaire créée avec {documents_count} document(s)",
        }
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur création ligne: {e}")
        raise HTTPException(500, f"Erreur lors de la création: {e!s}")


# ============================================
# ÉDITION DES ÉLÉMENTS HIÉRARCHIQUES
# ============================================


@router.put("/api/actions/{action_id}")
def api_update_action(
    action_id: int,
    libelle: str = Form(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Modifier le libellé d'une action (les montants sont recalculés automatiquement)"""
    action = session.get(ActionBudgetaire, action_id)
    if not action:
        raise HTTPException(404, "Action non trouvée")

    try:
        action.libelle = libelle
        action.updated_at = datetime.utcnow()
        session.add(action)
        session.commit()

        logger.info(f"✅ Action {action_id} modifiée par {current_user.email}")
        return {"ok": True, "message": "Action modifiée avec succès"}
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur modification action {action_id}: {e}")
        raise HTTPException(500, f"Erreur lors de la modification: {e!s}")


@router.put("/api/services/{service_id}")
def api_update_service(
    service_id: int,
    libelle: str = Form(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Modifier le libellé d'un service (les montants sont recalculés automatiquement)"""
    service = session.get(ServiceBeneficiaire, service_id)
    if not service:
        raise HTTPException(404, "Service non trouvé")

    try:
        service.libelle = libelle
        service.updated_at = datetime.utcnow()
        session.add(service)
        session.commit()

        logger.info(f"✅ Service {service_id} modifié par {current_user.email}")
        return {"ok": True, "message": "Service modifié avec succès"}
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur modification service {service_id}: {e}")
        raise HTTPException(500, f"Erreur lors de la modification: {e!s}")


@router.put("/api/activites/{activite_id}")
def api_update_activite(
    activite_id: int,
    libelle: str = Form(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Modifier le libellé d'une activité (les montants sont recalculés automatiquement)"""
    activite = session.get(ActiviteBudgetaire, activite_id)
    if not activite:
        raise HTTPException(404, "Activité non trouvée")

    try:
        activite.libelle = libelle
        activite.updated_at = datetime.utcnow()
        session.add(activite)
        session.commit()

        logger.info(f"✅ Activité {activite_id} modifiée par {current_user.email}")
        return {"ok": True, "message": "Activité modifiée avec succès"}
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur modification activité {activite_id}: {e}")
        raise HTTPException(500, f"Erreur lors de la modification: {e!s}")


@router.put("/api/lignes/{ligne_id}")
def api_update_ligne(
    ligne_id: int,
    libelle: str = Form(...),
    budget_vote_n: float | None = Form(None),
    budget_actuel_n: float | None = Form(None),
    enveloppe_n_plus_1: float | None = Form(None),
    complement_solicite: float | None = Form(None),
    engagement_etat: float | None = Form(None),
    autre_complement: float | None = Form(None),
    justificatifs: str | None = Form(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Modifier une ligne budgétaire (libellé + montants)"""
    ligne = session.get(LigneBudgetaireDetail, ligne_id)
    if not ligne:
        raise HTTPException(404, "Ligne budgétaire non trouvée")

    try:
        # Mise à jour des champs
        ligne.libelle = libelle
        if budget_vote_n is not None:
            ligne.budget_vote_n = Decimal(str(budget_vote_n))
        if budget_actuel_n is not None:
            ligne.budget_actuel_n = Decimal(str(budget_actuel_n))
        if enveloppe_n_plus_1 is not None:
            ligne.enveloppe_n_plus_1 = Decimal(str(enveloppe_n_plus_1))
        if complement_solicite is not None:
            ligne.complement_solicite = Decimal(str(complement_solicite))
        if engagement_etat is not None:
            ligne.engagement_etat = Decimal(str(engagement_etat))
        if autre_complement is not None:
            ligne.autre_complement = Decimal(str(autre_complement))

        # Recalculer les champs dérivés
        ligne.budget_souhaite = ligne.enveloppe_n_plus_1 + ligne.complement_solicite
        ligne.projet_budget_n_plus_1 = ligne.enveloppe_n_plus_1 + ligne.engagement_etat + ligne.autre_complement

        if justificatifs is not None:
            ligne.justificatifs = justificatifs

        ligne.updated_at = datetime.utcnow()
        session.add(ligne)
        session.commit()

        # Recalculer les totaux de toute la hiérarchie
        FicheTechniqueService._recalculer_totaux_hierarchie(ligne.fiche_technique_id, session)

        logger.info(f"✅ Ligne budgétaire {ligne_id} modifiée par {current_user.email}")
        return {"ok": True, "message": "Ligne budgétaire modifiée avec succès"}
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur modification ligne {ligne_id}: {e}")
        raise HTTPException(500, f"Erreur lors de la modification: {e!s}")


# ============================================
# SUPPRESSION DES ÉLÉMENTS HIÉRARCHIQUES
# ============================================


@router.delete("/api/actions/{action_id}")
def api_delete_action(
    action_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """Supprimer une action (uniquement si elle n'a pas d'enfants)"""
    action = session.get(ActionBudgetaire, action_id)
    if not action:
        raise HTTPException(404, "Action non trouvée")

    try:
        # Vérifier qu'il n'y a pas de services enfants
        services_count = session.exec(
            select(func.count(ServiceBeneficiaire.id)).where(ServiceBeneficiaire.action_id == action_id)
        ).one()

        if services_count > 0:
            raise HTTPException(
                400,
                f"Impossible de supprimer cette action : elle contient {services_count} service(s). Veuillez d'abord supprimer tous les services enfants.",
            )

        fiche_id = action.fiche_technique_id

        # Supprimer l'action (aucun enfant)
        session.delete(action)
        session.commit()

        # Recalculer les totaux
        FicheTechniqueService._recalculer_totaux_hierarchie(fiche_id, session)

        logger.info(f"✅ Action {action_id} supprimée par {current_user.email}")
        return {"ok": True, "message": "Action supprimée avec succès"}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur suppression action {action_id}: {e}")
        raise HTTPException(500, f"Erreur lors de la suppression: {e!s}")


@router.delete("/api/services/{service_id}")
def api_delete_service(
    service_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """Supprimer un service (uniquement s'il n'a pas d'enfants)"""
    service = session.get(ServiceBeneficiaire, service_id)
    if not service:
        raise HTTPException(404, "Service non trouvé")

    try:
        # Vérifier qu'il n'y a pas d'activités enfants
        activites_count = session.exec(
            select(func.count(ActiviteBudgetaire.id)).where(ActiviteBudgetaire.service_beneficiaire_id == service_id)
        ).one()

        if activites_count > 0:
            raise HTTPException(
                400,
                f"Impossible de supprimer ce service : il contient {activites_count} activité(s). Veuillez d'abord supprimer toutes les activités enfants.",
            )

        fiche_id = service.fiche_technique_id

        # Supprimer le service (aucun enfant)
        session.delete(service)
        session.commit()

        # Recalculer les totaux
        FicheTechniqueService._recalculer_totaux_hierarchie(fiche_id, session)

        logger.info(f"✅ Service {service_id} supprimé par {current_user.email}")
        return {"ok": True, "message": "Service supprimé avec succès"}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur suppression service {service_id}: {e}")
        raise HTTPException(500, f"Erreur lors de la suppression: {e!s}")


@router.delete("/api/activites/{activite_id}")
def api_delete_activite(
    activite_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """Supprimer une activité (uniquement si elle n'a pas de lignes)"""
    activite = session.get(ActiviteBudgetaire, activite_id)
    if not activite:
        raise HTTPException(404, "Activité non trouvée")

    try:
        # Vérifier qu'il n'y a pas de lignes budgétaires enfants
        lignes_count = session.exec(
            select(func.count(LigneBudgetaireDetail.id)).where(LigneBudgetaireDetail.activite_id == activite_id)
        ).one()

        if lignes_count > 0:
            raise HTTPException(
                400,
                f"Impossible de supprimer cette activité : elle contient {lignes_count} ligne(s) budgétaire(s). Veuillez d'abord supprimer toutes les lignes enfants.",
            )

        fiche_id = activite.fiche_technique_id

        # Supprimer l'activité (aucune ligne)
        session.delete(activite)
        session.commit()

        # Recalculer les totaux
        FicheTechniqueService._recalculer_totaux_hierarchie(fiche_id, session)

        logger.info(f"✅ Activité {activite_id} supprimée par {current_user.email}")
        return {"ok": True, "message": "Activité supprimée avec succès"}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur suppression activité {activite_id}: {e}")
        raise HTTPException(500, f"Erreur lors de la suppression: {e!s}")


@router.delete("/api/lignes/{ligne_id}")
def api_delete_ligne(
    ligne_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """Supprimer une ligne budgétaire et ses documents"""
    ligne = session.get(LigneBudgetaireDetail, ligne_id)
    if not ligne:
        raise HTTPException(404, "Ligne budgétaire non trouvée")

    try:
        fiche_id = ligne.fiche_technique_id

        # Supprimer les documents associés
        documents = session.exec(
            select(DocumentLigneBudgetaire).where(DocumentLigneBudgetaire.ligne_budgetaire_id == ligne_id)
        ).all()

        for doc in documents:
            # Supprimer le fichier physique
            try:
                file_path = Path(doc.chemin_fichier)
                if file_path.exists():
                    file_path.unlink()
                    logger.info(f"📎 Fichier supprimé : {doc.nom_fichier_stocke}")
            except Exception as e:
                logger.warning(f"⚠️ Impossible de supprimer le fichier {doc.chemin_fichier}: {e}")

            # Supprimer la métadonnée
            session.delete(doc)

        # Supprimer la ligne
        session.delete(ligne)
        session.commit()

        # Recalculer les totaux
        FicheTechniqueService._recalculer_totaux_hierarchie(fiche_id, session)

        logger.info(f"✅ Ligne budgétaire {ligne_id} et ses documents supprimés par {current_user.email}")
        return {"ok": True, "message": "Ligne budgétaire supprimée avec succès"}
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur suppression ligne {ligne_id}: {e}")
        raise HTTPException(500, f"Erreur lors de la suppression: {e!s}")


# ============================================
# GESTION DES DOCUMENTS DES LIGNES
# ============================================


@router.get("/api/lignes/{ligne_id}/documents")
def api_get_documents_ligne(
    ligne_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """Récupérer la liste des documents d'une ligne budgétaire"""
    ligne = session.get(LigneBudgetaireDetail, ligne_id)
    if not ligne:
        raise HTTPException(404, "Ligne budgétaire non trouvée")

    documents = session.exec(
        select(DocumentLigneBudgetaire)
        .where(DocumentLigneBudgetaire.ligne_budgetaire_id == ligne_id)
        .where(DocumentLigneBudgetaire.actif)
        .order_by(DocumentLigneBudgetaire.uploaded_at.desc())
    ).all()

    from app.core.path_config import path_config

    return {
        "ok": True,
        "documents": [
            {
                "id": doc.id,
                "nom_original": doc.nom_fichier_original,
                "nom_stocke": doc.nom_fichier_stocke,
                "type": doc.type_fichier,
                "taille": doc.taille_octets,
                "url": path_config.get_file_url("uploads", f"budget/lignes/{ligne_id}/{doc.nom_fichier_stocke}"),
                "uploaded_at": doc.uploaded_at.isoformat(),
            }
            for doc in documents
        ],
    }


@router.post("/api/lignes/{ligne_id}/documents")
async def api_add_documents_ligne(
    ligne_id: int,
    documents: list[UploadFile] = File(...),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Ajouter des documents à une ligne budgétaire existante"""
    ligne = session.get(LigneBudgetaireDetail, ligne_id)
    if not ligne:
        raise HTTPException(404, "Ligne budgétaire non trouvée")

    try:
        # Récupérer les codes pour le renommage
        activite = session.get(ActiviteBudgetaire, ligne.activite_id)
        service = session.get(ServiceBeneficiaire, activite.service_beneficiaire_id)
        action = session.get(ActionBudgetaire, service.action_id)

        docs_dir = Path(f"uploads/budget/lignes/{ligne_id}")
        docs_dir.mkdir(parents=True, exist_ok=True)

        documents_count = 0
        for doc in documents:
            if doc.filename:
                # Renommer le fichier selon le format
                file_ext = Path(doc.filename).suffix
                original_name = Path(doc.filename).stem
                new_filename = f"{action.code}_{activite.code}_{ligne.code}_{original_name}{file_ext}"

                # Sauvegarder le fichier
                file_path = docs_dir / new_filename
                content = await doc.read()

                with open(file_path, "wb") as f:
                    f.write(content)

                # Enregistrer les métadonnées
                doc_meta = DocumentLigneBudgetaire(
                    ligne_budgetaire_id=ligne_id,
                    fiche_technique_id=ligne.fiche_technique_id,
                    nom_fichier_original=doc.filename,
                    nom_fichier_stocke=new_filename,
                    chemin_fichier=str(file_path),
                    type_fichier=file_ext,
                    taille_octets=len(content),
                    code_action=action.code,
                    code_activite=activite.code,
                    code_ligne=ligne.code,
                    uploaded_by_user_id=current_user.id,
                )
                session.add(doc_meta)
                documents_count += 1
                logger.info(f"📎 Document ajouté : {new_filename}")

        session.commit()

        logger.info(f"✅ {documents_count} document(s) ajouté(s) à la ligne {ligne_id}")
        return {
            "ok": True,
            "documents_count": documents_count,
            "message": f"{documents_count} document(s) ajouté(s) avec succès",
        }
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur ajout documents: {e}")
        raise HTTPException(500, f"Erreur lors de l'ajout: {e!s}")


@router.delete("/api/lignes/{ligne_id}/documents/{document_id}")
def api_delete_document_ligne(
    ligne_id: int,
    document_id: int,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Supprimer un document d'une ligne budgétaire"""
    doc = session.get(DocumentLigneBudgetaire, document_id)
    if not doc or doc.ligne_budgetaire_id != ligne_id:
        raise HTTPException(404, "Document non trouvé")

    try:
        # Supprimer le fichier physique
        file_path = Path(doc.chemin_fichier)
        if file_path.exists():
            file_path.unlink()
            logger.info(f"📎 Fichier supprimé : {doc.nom_fichier_stocke}")

        # Supprimer la métadonnée
        session.delete(doc)
        session.commit()

        logger.info(f"✅ Document {document_id} supprimé par {current_user.email}")
        return {"ok": True, "message": "Document supprimé avec succès"}
    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur suppression document {document_id}: {e}")
        raise HTTPException(500, f"Erreur lors de la suppression: {e!s}")


# ============================================
# SIGOBE - Système d'Information de Gestion et d'Observation Budgétaire
# ============================================

# --- Helpers de parsing SIGOBE (inspirés de PowerQuery) ---


def remove_accents(text: str) -> str:
    """Supprime les accents d'un texte"""
    if not text:
        return ""

    replacements = {
        "à": "a",
        "á": "a",
        "â": "a",
        "ä": "a",
        "ã": "a",
        "å": "a",
        "è": "e",
        "é": "e",
        "ê": "e",
        "ë": "e",
        "ì": "i",
        "í": "i",
        "î": "i",
        "ï": "i",
        "ò": "o",
        "ó": "o",
        "ô": "o",
        "ö": "o",
        "õ": "o",
        "ù": "u",
        "ú": "u",
        "û": "u",
        "ü": "u",
        "ç": "c",
        "ñ": "n",
        "À": "A",
        "Á": "A",
        "Â": "A",
        "Ä": "A",
        "Ã": "A",
        "Å": "A",
        "È": "E",
        "É": "E",
        "Ê": "E",
        "Ë": "E",
        "Ì": "I",
        "Í": "I",
        "Î": "I",
        "Ï": "I",
        "Ò": "O",
        "Ó": "O",
        "Ô": "O",
        "Ö": "O",
        "Õ": "O",
        "Ù": "U",
        "Ú": "U",
        "Û": "U",
        "Ü": "U",
        "Ç": "C",
        "Ñ": "N",
    }

    for accent, replacement in replacements.items():
        text = text.replace(accent, replacement)

    # Remplacer nbsp par espace
    text = text.replace("\u00a0", " ")

    return text


def normalize_text(text: str) -> str:
    """Normalise un texte : minuscules, sans accents, trim"""
    if not text or pd.isna(text):
        return ""
    text = str(text).strip()
    text = remove_accents(text)
    text = text.lower()
    return text


def parse_date_flexible(value) -> date | None:
    """Parse une date de manière flexible (texte, nombre Excel, date)"""
    if pd.isna(value) or value is None:
        return None

    # Si déjà une date
    if isinstance(value, (date, datetime)):
        return value if isinstance(value, date) else value.date()

    # Si nombre (date Excel)
    try:
        if isinstance(value, (int, float)):
            return pd.to_datetime(value, origin="1899-12-30", unit="D").date()
    except:
        pass

    # Si texte
    try:
        text = str(value).strip().replace(".", "/").replace("-", "/")
        parsed = pd.to_datetime(text, format="%d/%m/%Y", errors="coerce")
        if pd.notna(parsed):
            return parsed.date()
    except:
        pass

    return None


def standardize_column_names(df: pd.DataFrame) -> pd.DataFrame:
    """
    Renomme automatiquement les colonnes financières
    Inspiré de fxTableStandardName_AutoMapping
    ⚠️ ORDRE IMPORTANT : Du plus spécifique au plus général !
    """
    # Ordre crucial : patterns spécifiques en premier
    standard_mapping = [
        # Mandats (spécifiques en premier !)
        (["vise cf", "visé cf", "vise"], "Mandats_Vise_CF"),
        (["pec", "prise en charge"], "Mandats_Pec"),
        (["mandat emis", "mandats emis"], "Mandats_Emis"),  # Spécifique
        (["emis", "emise"], "Mandats_Emis"),  # Général (après les spécifiques)
        # Budget
        (["budget vote", "budget voté"], "Budget_Vote"),
        (["budget actuel", "actuel"], "Budget_Actuel"),
        (["vote"], "Budget_Vote"),
        (["actuel"], "Budget_Actuel"),
        # Engagements
        (["engagement", "engagements", "engag"], "Engagements_Emis"),
        # Disponible
        (["disponible", "dispo"], "Disponible_Eng"),
    ]

    new_columns = []
    for col in df.columns:
        col_normalized = normalize_text(str(col))

        # Chercher une correspondance (du plus spécifique au plus général)
        matched = False
        for patterns, std_name in standard_mapping:
            for pattern in patterns:
                if pattern in col_normalized:
                    new_columns.append(std_name)
                    matched = True
                    break
            if matched:
                break

        if not matched:
            new_columns.append(col)

    df.columns = new_columns
    return df


def split_code_libelle(text: str) -> tuple:
    """
    Sépare code et libellé (ex: '2208401 Pilotage...' -> ('2208401', 'Pilotage...')).
    Le code doit être numérique, sinon pas de code.

    Exemples:
    - "2208401 Pilotage et Soutien" -> ("2208401", "Pilotage et Soutien")
    - "P01 Programme 01" -> ("P01", "Programme 01")
    - "Direction Générale" -> ("", "Direction Générale")
    - "Pilotage" -> ("", "Pilotage")
    """
    if not text or pd.isna(text):
        return ("", "")

    text = str(text).strip()
    parts = text.split(" ", 1)

    if len(parts) == 2:
        first_part = parts[0].strip()
        # Vérifier si le premier mot contient au moins un chiffre (code numérique)
        if any(c.isdigit() for c in first_part):
            # C'est un code numérique suivi d'un libellé
            return (first_part, parts[1].strip())
        else:
            # Pas de code numérique, tout est le libellé
            return ("", text)
    else:
        # Un seul mot, pas de séparation -> pas de code
        return ("", text)


@router.get("/sigobe", response_class=HTMLResponse, name="budget_sigobe")
def budget_sigobe(
    request: Request,
    annee: int | None = None,
    trimestre: int | None = None,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Page principale SIGOBE"""

    # Récupérer les chargements
    query = select(SigobeChargement).order_by(SigobeChargement.date_chargement.desc())

    if annee:
        query = query.where(SigobeChargement.annee == annee)
    if trimestre:
        query = query.where(SigobeChargement.trimestre == trimestre)

    chargements = session.exec(query).all()

    # Années disponibles
    annees_query = select(SigobeChargement.annee).distinct()
    annees_result = session.exec(annees_query).all()
    # Gérer le cas où c'est des int directs ou des tuples
    if annees_result and isinstance(annees_result[0], tuple):
        annees = sorted([a for (a,) in annees_result], reverse=True)
    else:
        annees = sorted(annees_result, reverse=True)

    # KPIs globaux (dernière période)
    kpis_global = None
    if chargements:
        dernier_chargement = chargements[0]
        kpis_global = session.exec(
            select(SigobeKpi)
            .where(SigobeKpi.chargement_id == dernier_chargement.id)
            .where(SigobeKpi.dimension == "global")
        ).first()

    return templates.TemplateResponse(
        "pages/budget_sigobe.html",
        get_template_context(
            request,
            chargements=chargements,
            annees=annees,
            annee_selected=annee,
            trimestre_selected=trimestre,
            kpis_global=kpis_global,
            current_user=current_user,
        ),
    )


@router.get("/sigobe/{chargement_id}/table", response_class=HTMLResponse)
def budget_sigobe_table(
    chargement_id: int,
    request: Request,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Page de détail d'un chargement SIGOBE avec table éditable"""

    # Récupérer le chargement
    chargement = session.get(SigobeChargement, chargement_id)
    if not chargement:
        raise HTTPException(404, "Chargement SIGOBE non trouvé")

    # Récupérer les lignes d'exécution
    executions = session.exec(
        select(SigobeExecution)
        .where(SigobeExecution.chargement_id == chargement_id)
        .order_by(
            SigobeExecution.programmes, SigobeExecution.actions, SigobeExecution.activites, SigobeExecution.taches
        )
    ).all()

    # Extraire les valeurs uniques pour les filtres
    programmes = sorted(set(e.programmes for e in executions if e.programmes))
    actions = sorted(set(e.actions for e in executions if e.actions))
    types_depense = sorted(set(e.type_depense for e in executions if e.type_depense))

    return templates.TemplateResponse(
        "pages/budget_sigobe_table.html",
        get_template_context(
            request,
            chargement=chargement,
            executions=executions,
            programmes=programmes,
            actions=actions,
            types_depense=types_depense,
        ),
    )


@router.get("/api/telecharger-template-sigobe")
async def api_telecharger_template_sigobe(current_user: User = Depends(get_current_user)):
    """
    Télécharger un modèle Excel vierge pour les données SIGOBE
    """
    try:
        # Créer un classeur Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "SIGOBE - Situation Exécution"

        # Définir les styles
        header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")  # Orange
        header_font = Font(bold=True, color="FFFFFF", size=11)
        example_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        total_action_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")  # Bleu clair
        total_programme_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")  # Or
        total_general_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")  # Rouge
        total_font = Font(bold=True, size=10)
        total_general_font = Font(bold=True, size=11, color="FFFFFF")
        border_style = Border(
            left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
        )

        # Définir les en-têtes de colonnes SIGOBE (hiérarchie + finances)
        headers = [
            "PROGRAMMES",
            "ACTIONS",
            "RPROG",
            "TYPE DEPENSE",
            "ACTIVITES",
            "TACHES",
            "BUDGET VOTE",
            "BUDGET ACTUEL",
            "ENGAGEMENTS EMIS",
            "DISPONIBLE ENG",
            "MANDATS EMIS",
            "MANDATS VISE CF",
            "MANDATS PEC",
        ]

        # Écrire les en-têtes (ligne 1 directement, sans titre pour éviter confusion au parsing)
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border_style

        # Ajuster la largeur des colonnes
        # Colonnes hiérarchiques (A-F) plus larges
        for col in ["A", "B", "C", "D", "E", "F"]:
            ws.column_dimensions[col].width = 30
        # Colonnes financières (G-M)
        for col in ["G", "H", "I", "J", "K", "L", "M"]:
            ws.column_dimensions[col].width = 18

        # Ajouter des exemples avec la hiérarchie complète (13 colonnes)
        # Format: (PROGRAMMES, ACTIONS, RPROG, TYPE DEPENSE, ACTIVITES, TACHES, BUDGET VOTE, BUDGET ACTUEL, ENGAGEMENTS EMIS, DISPONIBLE ENG, MANDATS EMIS, MANDATS VISE CF, MANDATS PEC)
        exemples = [
            # ===== PROGRAMME 001 : PILOTAGE ET COORDINATION =====
            ("Programme 001 - Pilotage et Coordination", "", "", "", "", "", "", "", "", "", "", "", ""),
            # Action 1.1
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.1 - Coordination administrative",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.1 - Coordination administrative",
                "RPROG-001",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.1 - Coordination administrative",
                "RPROG-001",
                "Fonctionnement",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.1 - Coordination administrative",
                "RPROG-001",
                "Fonctionnement",
                "Activité 1.1.1 - Gestion courante",
                "",
                "50000000",
                "48000000",
                "35000000",
                "13000000",
                "30000000",
                "28000000",
                "25000000",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.1 - Coordination administrative",
                "RPROG-001",
                "Fonctionnement",
                "Activité 1.1.1 - Gestion courante",
                "Tâche 1.1.1.1 - Fournitures de bureau",
                "10000000",
                "9500000",
                "7000000",
                "2500000",
                "6000000",
                "5500000",
                "5000000",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.1 - Coordination administrative",
                "RPROG-001",
                "Fonctionnement",
                "Activité 1.1.1 - Gestion courante",
                "Tâche 1.1.1.2 - Matériel informatique",
                "15000000",
                "14500000",
                "10000000",
                "4500000",
                "8500000",
                "8000000",
                "7500000",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.1 - Coordination administrative",
                "RPROG-001",
                "Fonctionnement",
                "Activité 1.1.1 - Gestion courante",
                "Tâche 1.1.1.3 - Entretien des locaux",
                "25000000",
                "24000000",
                "18000000",
                "6000000",
                "16000000",
                "14500000",
                "12500000",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.1 - Coordination administrative",
                "RPROG-001",
                "Fonctionnement",
                "Activité 1.1.2 - Communication institutionnelle",
                "",
                "30000000",
                "29000000",
                "20000000",
                "9000000",
                "18000000",
                "17000000",
                "15000000",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.1 - Coordination administrative",
                "RPROG-001",
                "Fonctionnement",
                "Activité 1.1.2 - Communication institutionnelle",
                "Tâche 1.1.2.1 - Supports de communication",
                "12000000",
                "11500000",
                "8000000",
                "3500000",
                "7000000",
                "6500000",
                "6000000",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.1 - Coordination administrative",
                "RPROG-001",
                "Fonctionnement",
                "Activité 1.1.2 - Communication institutionnelle",
                "Tâche 1.1.2.2 - Événements institutionnels",
                "18000000",
                "17500000",
                "12000000",
                "5500000",
                "11000000",
                "10500000",
                "9000000",
            ),
            # TOTAL Action 1.1
            (
                "Programme 001 - Pilotage et Coordination",
                "TOTAL Action 1.1 - Coordination administrative",
                "",
                "",
                "",
                "",
                "80000000",
                "77000000",
                "55000000",
                "22000000",
                "48000000",
                "45000000",
                "40000000",
            ),
            # Action 1.2
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.2 - Suivi et évaluation",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.2 - Suivi et évaluation",
                "RPROG-001",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.2 - Suivi et évaluation",
                "RPROG-001",
                "Fonctionnement",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.2 - Suivi et évaluation",
                "RPROG-001",
                "Fonctionnement",
                "Activité 1.2.1 - Suivi des performances",
                "",
                "20000000",
                "19500000",
                "14000000",
                "5500000",
                "12000000",
                "11500000",
                "10000000",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.2 - Suivi et évaluation",
                "RPROG-001",
                "Fonctionnement",
                "Activité 1.2.1 - Suivi des performances",
                "Tâche 1.2.1.1 - Indicateurs de performance",
                "8000000",
                "7800000",
                "5500000",
                "2300000",
                "5000000",
                "4800000",
                "4200000",
            ),
            (
                "Programme 001 - Pilotage et Coordination",
                "Action 1.2 - Suivi et évaluation",
                "RPROG-001",
                "Fonctionnement",
                "Activité 1.2.1 - Suivi des performances",
                "Tâche 1.2.1.2 - Rapports d'activité",
                "12000000",
                "11700000",
                "8500000",
                "3200000",
                "7000000",
                "6700000",
                "5800000",
            ),
            # TOTAL Action 1.2
            (
                "Programme 001 - Pilotage et Coordination",
                "TOTAL Action 1.2 - Suivi et évaluation",
                "",
                "",
                "",
                "",
                "20000000",
                "19500000",
                "14000000",
                "5500000",
                "12000000",
                "11500000",
                "10000000",
            ),
            # TOTAL PROGRAMME 001
            (
                "TOTAL Programme 001 - Pilotage et Coordination",
                "",
                "",
                "",
                "",
                "",
                "100000000",
                "96500000",
                "69000000",
                "27500000",
                "60000000",
                "57500000",
                "50000000",
            ),
            # ===== PROGRAMME 002 : DÉVELOPPEMENT STRATÉGIQUE =====
            ("Programme 002 - Développement Stratégique", "", "", "", "", "", "", "", "", "", "", "", ""),
            # Action 2.1
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.1 - Études et planification",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.1 - Études et planification",
                "RPROG-002",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.1 - Études et planification",
                "RPROG-002",
                "Fonctionnement",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.1 - Études et planification",
                "RPROG-002",
                "Fonctionnement",
                "Activité 2.1.1 - Analyses sectorielles",
                "",
                "25000000",
                "24000000",
                "18000000",
                "6000000",
                "15000000",
                "14000000",
                "13000000",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.1 - Études et planification",
                "RPROG-002",
                "Fonctionnement",
                "Activité 2.1.1 - Analyses sectorielles",
                "Tâche 2.1.1.1 - Études de marché",
                "12000000",
                "11500000",
                "9000000",
                "2500000",
                "7500000",
                "7000000",
                "6500000",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.1 - Études et planification",
                "RPROG-002",
                "Fonctionnement",
                "Activité 2.1.1 - Analyses sectorielles",
                "Tâche 2.1.1.2 - Consultations externes",
                "13000000",
                "12500000",
                "9000000",
                "3500000",
                "7500000",
                "7000000",
                "6500000",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.1 - Études et planification",
                "RPROG-002",
                "Fonctionnement",
                "Activité 2.1.2 - Planification stratégique",
                "",
                "35000000",
                "34000000",
                "25000000",
                "9000000",
                "22000000",
                "20000000",
                "18000000",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.1 - Études et planification",
                "RPROG-002",
                "Fonctionnement",
                "Activité 2.1.2 - Planification stratégique",
                "Tâche 2.1.2.1 - Élaboration des plans",
                "20000000",
                "19500000",
                "14000000",
                "5500000",
                "12000000",
                "11000000",
                "10000000",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.1 - Études et planification",
                "RPROG-002",
                "Fonctionnement",
                "Activité 2.1.2 - Planification stratégique",
                "Tâche 2.1.2.2 - Ateliers de validation",
                "15000000",
                "14500000",
                "11000000",
                "3500000",
                "10000000",
                "9000000",
                "8000000",
            ),
            # TOTAL Action 2.1
            (
                "Programme 002 - Développement Stratégique",
                "TOTAL Action 2.1 - Études et planification",
                "",
                "",
                "",
                "",
                "60000000",
                "58000000",
                "43000000",
                "15000000",
                "37000000",
                "34000000",
                "31000000",
            ),
            # Action 2.2
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.2 - Mise en œuvre des projets",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.2 - Mise en œuvre des projets",
                "RPROG-002",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.2 - Mise en œuvre des projets",
                "RPROG-002",
                "Investissement",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.2 - Mise en œuvre des projets",
                "RPROG-002",
                "Investissement",
                "Activité 2.2.1 - Infrastructures",
                "",
                "80000000",
                "78000000",
                "55000000",
                "23000000",
                "50000000",
                "48000000",
                "45000000",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.2 - Mise en œuvre des projets",
                "RPROG-002",
                "Investissement",
                "Activité 2.2.1 - Infrastructures",
                "Tâche 2.2.1.1 - Construction bâtiments",
                "50000000",
                "49000000",
                "35000000",
                "14000000",
                "32000000",
                "31000000",
                "29000000",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.2 - Mise en œuvre des projets",
                "RPROG-002",
                "Investissement",
                "Activité 2.2.1 - Infrastructures",
                "Tâche 2.2.1.2 - Équipements techniques",
                "30000000",
                "29000000",
                "20000000",
                "9000000",
                "18000000",
                "17000000",
                "16000000",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.2 - Mise en œuvre des projets",
                "RPROG-002",
                "Investissement",
                "Activité 2.2.2 - Équipements informatiques",
                "",
                "45000000",
                "44000000",
                "32000000",
                "12000000",
                "28000000",
                "27000000",
                "25000000",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.2 - Mise en œuvre des projets",
                "RPROG-002",
                "Investissement",
                "Activité 2.2.2 - Équipements informatiques",
                "Tâche 2.2.2.1 - Serveurs et réseaux",
                "25000000",
                "24500000",
                "18000000",
                "6500000",
                "16000000",
                "15500000",
                "14500000",
            ),
            (
                "Programme 002 - Développement Stratégique",
                "Action 2.2 - Mise en œuvre des projets",
                "RPROG-002",
                "Investissement",
                "Activité 2.2.2 - Équipements informatiques",
                "Tâche 2.2.2.2 - Postes de travail",
                "20000000",
                "19500000",
                "14000000",
                "5500000",
                "12000000",
                "11500000",
                "10500000",
            ),
            # TOTAL Action 2.2
            (
                "Programme 002 - Développement Stratégique",
                "TOTAL Action 2.2 - Mise en œuvre des projets",
                "",
                "",
                "",
                "",
                "125000000",
                "122000000",
                "87000000",
                "35000000",
                "78000000",
                "75000000",
                "70000000",
            ),
            # TOTAL PROGRAMME 002
            (
                "TOTAL Programme 002 - Développement Stratégique",
                "",
                "",
                "",
                "",
                "",
                "185000000",
                "180000000",
                "130000000",
                "50000000",
                "115000000",
                "109000000",
                "101000000",
            ),
            # ===== PROGRAMME 003 : RESSOURCES HUMAINES =====
            ("Programme 003 - Gestion des Ressources Humaines", "", "", "", "", "", "", "", "", "", "", "", ""),
            # Action 3.1
            (
                "Programme 003 - Gestion des Ressources Humaines",
                "Action 3.1 - Formation et développement",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 003 - Gestion des Ressources Humaines",
                "Action 3.1 - Formation et développement",
                "RPROG-003",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 003 - Gestion des Ressources Humaines",
                "Action 3.1 - Formation et développement",
                "RPROG-003",
                "Fonctionnement",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
            ),
            (
                "Programme 003 - Gestion des Ressources Humaines",
                "Action 3.1 - Formation et développement",
                "RPROG-003",
                "Fonctionnement",
                "Activité 3.1.1 - Formation continue",
                "",
                "40000000",
                "38500000",
                "28000000",
                "10500000",
                "25000000",
                "24000000",
                "22000000",
            ),
            (
                "Programme 003 - Gestion des Ressources Humaines",
                "Action 3.1 - Formation et développement",
                "RPROG-003",
                "Fonctionnement",
                "Activité 3.1.1 - Formation continue",
                "Tâche 3.1.1.1 - Formations techniques",
                "25000000",
                "24000000",
                "17500000",
                "6500000",
                "15000000",
                "14500000",
                "13500000",
            ),
            (
                "Programme 003 - Gestion des Ressources Humaines",
                "Action 3.1 - Formation et développement",
                "RPROG-003",
                "Fonctionnement",
                "Activité 3.1.1 - Formation continue",
                "Tâche 3.1.1.2 - Formations managériales",
                "15000000",
                "14500000",
                "10500000",
                "4000000",
                "10000000",
                "9500000",
                "8500000",
            ),
            (
                "Programme 003 - Gestion des Ressources Humaines",
                "Action 3.1 - Formation et développement",
                "RPROG-003",
                "Fonctionnement",
                "Activité 3.1.2 - Recrutement",
                "",
                "15000000",
                "14800000",
                "10000000",
                "4800000",
                "8500000",
                "8200000",
                "7500000",
            ),
            (
                "Programme 003 - Gestion des Ressources Humaines",
                "Action 3.1 - Formation et développement",
                "RPROG-003",
                "Fonctionnement",
                "Activité 3.1.2 - Recrutement",
                "Tâche 3.1.2.1 - Procédures de sélection",
                "10000000",
                "9800000",
                "7000000",
                "2800000",
                "6000000",
                "5800000",
                "5300000",
            ),
            (
                "Programme 003 - Gestion des Ressources Humaines",
                "Action 3.1 - Formation et développement",
                "RPROG-003",
                "Fonctionnement",
                "Activité 3.1.2 - Recrutement",
                "Tâche 3.1.2.2 - Tests et évaluations",
                "5000000",
                "5000000",
                "3000000",
                "2000000",
                "2500000",
                "2400000",
                "2200000",
            ),
            # TOTAL Action 3.1
            (
                "Programme 003 - Gestion des Ressources Humaines",
                "TOTAL Action 3.1 - Formation et développement",
                "",
                "",
                "",
                "",
                "55000000",
                "53300000",
                "38000000",
                "15300000",
                "33500000",
                "32200000",
                "29500000",
            ),
            # TOTAL PROGRAMME 003
            (
                "TOTAL Programme 003 - Gestion des Ressources Humaines",
                "",
                "",
                "",
                "",
                "",
                "55000000",
                "53300000",
                "38000000",
                "15300000",
                "33500000",
                "32200000",
                "29500000",
            ),
            # ===== TOTAL GÉNÉRAL =====
            ("", "", "", "", "", "", "", "", "", "", "", "", ""),
            (
                "TOTAL GÉNÉRAL - TOUS PROGRAMMES",
                "",
                "",
                "",
                "",
                "",
                "340000000",
                "329800000",
                "237000000",
                "92800000",
                "208500000",
                "198700000",
                "180500000",
            ),
        ]

        current_row = 2
        for exemple in exemples:
            # Déterminer le type de ligne pour le style
            first_value = str(exemple[0]) if exemple[0] else ""
            second_value = str(exemple[1]) if len(exemple) > 1 and exemple[1] else ""

            is_total_action = second_value.startswith("TOTAL Action")
            is_total_programme = first_value.startswith("TOTAL Programme")
            is_total_general = first_value.startswith("TOTAL GÉNÉRAL")

            for col_num, value in enumerate(exemple, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = value
                cell.border = border_style

                # Appliquer les styles selon le type de ligne
                if is_total_general:
                    cell.fill = total_general_fill
                    cell.font = total_general_font
                elif is_total_programme:
                    cell.fill = total_programme_fill
                    cell.font = total_font
                elif is_total_action:
                    cell.fill = total_action_fill
                    cell.font = total_font
                elif current_row > 2:
                    cell.fill = example_fill

                # Aligner les nombres à droite (colonnes financières G-M = 7-13)
                if col_num > 6 and value and value != "":
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.number_format = "#,##0"
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

            current_row += 1

        # Ajouter une feuille d'instructions
        ws_instructions = wb.create_sheet("📋 Instructions")
        ws_instructions.column_dimensions["A"].width = 80

        instructions = [
            ("MODÈLE SIGOBE - SITUATION D'EXÉCUTION BUDGÉTAIRE", header_font, header_fill),
            ("", None, None),
            ("📊 STRUCTURE HIÉRARCHIQUE (Colonnes A-F)", Font(bold=True, size=12), None),
            ("", None, None),
            ("1. PROGRAMMES : Libellé du programme budgétaire", None, example_fill),
            ("2. ACTIONS : Libellé de l'action (sous le programme)", None, example_fill),
            ("3. RPROG : Code du responsable de programme", None, example_fill),
            ("4. TYPE DEPENSE : Type de dépense (Fonctionnement, Investissement, etc.)", None, example_fill),
            ("5. ACTIVITES : Libellé de l'activité", None, example_fill),
            ("6. TACHES : Libellé de la tâche (niveau le plus détaillé)", None, example_fill),
            ("", None, None),
            ("💰 COLONNES FINANCIÈRES (Colonnes G-M)", Font(bold=True, size=12), None),
            ("", None, None),
            ("7. BUDGET VOTE : Montant du budget voté", None, None),
            ("8. BUDGET ACTUEL : Montant du budget actuel (après modifications)", None, None),
            ("9. ENGAGEMENTS EMIS : Montant des engagements émis", None, None),
            ("10. DISPONIBLE ENG : Montant disponible pour engagement", None, None),
            ("11. MANDATS EMIS : Montant des mandats émis", None, None),
            ("12. MANDATS VISE CF : Mandats visés par le contrôle financier", None, None),
            ("13. MANDATS PEC : Mandats pris en charge", None, None),
            ("", None, None),
            ("", None, None),
            ("⚠️ RÈGLES IMPORTANTES", Font(bold=True, size=12, color="DC3545"), None),
            ("", None, None),
            ("✅ RÉPÉTEZ la hiérarchie parente sur chaque ligne pour faciliter le tri/filtrage", None, None),
            ("✅ Chaque niveau hiérarchique a sa propre colonne (une colonne = un niveau)", None, None),
            ("✅ Sur une ligne, remplissez TOUS les niveaux parents + le niveau actuel", None, None),
            ("✅ Les montants doivent être des nombres entiers (sans espace ni symbole)", None, None),
            ("✅ Supprimez TOUTES les lignes d'exemples en gris avant l'import", None, None),
            ("", None, None),
            ("", None, None),
            ("💡 EXEMPLE DE HIÉRARCHIE COMPLÈTE", Font(bold=True, size=12, color="28A745"), None),
            ("", None, None),
            ("Pour une tâche au niveau le plus détaillé, RÉPÉTEZ tous les niveaux parents :", None, None),
            ("", None, None),
            ("Ligne Programme : [Programme 001] [vide] [vide] [vide] [vide] [vide] [montants...]", None, None),
            ("Ligne Action    : [Programme 001] [Action 1.1] [vide] [vide] [vide] [vide] [montants...]", None, None),
            (
                "Ligne RPROG     : [Programme 001] [Action 1.1] [RPROG-001] [vide] [vide] [vide] [montants...]",
                None,
                None,
            ),
            (
                "Ligne Type Dép. : [Programme 001] [Action 1.1] [RPROG-001] [Fonctionnement] [vide] [vide] [montants...]",
                None,
                None,
            ),
            (
                "Ligne Activité  : [Programme 001] [Action 1.1] [RPROG-001] [Fonctionnement] [Activité 1.1.1] [vide] [montants...]",
                None,
                None,
            ),
            (
                "Ligne Tâche     : [Programme 001] [Action 1.1] [RPROG-001] [Fonctionnement] [Activité 1.1.1] [Tâche 1.1.1.1] [montants...]",
                None,
                None,
            ),
            ("", None, None),
            ("⚡ AVANTAGES : Chaque ligne est autonome et peut être triée/filtrée facilement dans Excel", None, None),
            ("", None, None),
            ("", None, None),
            (f"📅 Modèle généré le {datetime.now().strftime('%d/%m/%Y à %H:%M')}", Font(italic=True, size=9), None),
        ]

        for row_num, (text, font, fill) in enumerate(instructions, 1):
            cell = ws_instructions.cell(row=row_num, column=1)
            cell.value = text
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Figer la première ligne (en-têtes uniquement, pas de titre)
        ws.freeze_panes = "A2"

        # Sauvegarder dans un buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        # Log de l'activité
        logger.info(f"📥 Template SIGOBE téléchargé par {current_user.email}")

        # Retourner le fichier
        filename = "Modele_SIGOBE.xlsx"
        return StreamingResponse(
            buffer,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"},
        )

    except Exception as e:
        logger.error(f"Erreur génération template SIGOBE: {e}")
        raise HTTPException(500, f"Erreur lors de la génération du modèle: {e!s}")


@router.post("/api/sigobe/preview")
async def api_sigobe_preview(
    fichier: UploadFile = File(...),
    annee: int = Form(...),
    trimestre: int | None = Form(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """
    Prévisualisation d'un fichier SIGOBE (sans sauvegarde)
    Analyse le fichier et retourne un aperçu des données
    """
    try:
        content = await fichier.read()
        excel_file = BytesIO(content)

        # Parser le fichier SIGOBE avec le service
        Result, Metadatafile, ColsToKeep = SigobeService.parse_fichier_excel(excel_file, annee, trimestre)

        # Statistiques
        stats = {
            "nb_lignes": len(Result),
            "nb_colonnes": len(Result.columns),
            "colonnes": list(Result.columns),
            "colonnes_hierarchiques": list(ColsToKeep),
            "nb_programmes": int(Result["Programmes"].nunique()) if "Programmes" in Result.columns else 0,
            "nb_actions": int(Result["Actions"].nunique()) if "Actions" in Result.columns else 0,
            "metadata": {k: str(v) for k, v in Metadatafile.items()},  # Convertir en strings
        }

        # Aperçu des données (20 premières lignes)
        preview_data = []
        for idx, row in Result.head(20).iterrows():
            preview_data.append({col: str(row[col]) if pd.notna(row[col]) else "" for col in Result.columns})

        # Calcul des totaux préliminaires
        financial_cols = ["Budget_Vote", "Budget_Actuel", "Engagements_Emis", "Mandats_Emis"]
        totaux = {}

        for col in financial_cols:
            if col in Result.columns:
                total = Result[col].sum()
                totaux[col] = float(total)

        logger.info(f"✅ Prévisualisation : {stats['nb_lignes']} lignes, {stats['nb_programmes']} programmes")

        return {
            "ok": True,
            "stats": stats,
            "preview": preview_data,
            "totaux": totaux,
            "message": f"Fichier analysé : {stats['nb_lignes']} lignes prêtes à importer",
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erreur prévisualisation SIGOBE : {e}")
        raise HTTPException(500, f"Erreur lors de l'analyse : {e!s}")


def _parse_sigobe_file(excel_file: BytesIO, annee: int, trimestre: int | None) -> tuple:
    """
    Parse un fichier SIGOBE - Wrapper vers SigobeService
    """
    return SigobeService.parse_fichier_excel(excel_file, annee, trimestre)


@router.post("/api/sigobe/upload")
async def api_sigobe_upload(
    fichier: UploadFile = File(...),
    annee: int = Form(...),
    trimestre: int | None = Form(None),
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """
    Upload et analyse d'un fichier SIGOBE (Excel)
    Parser suivant SCRUPULEUSEMENT la logique fxInspectTable PowerQuery (étapes A→O)
    """
    try:
        content = await fichier.read()
        excel_file = BytesIO(content)

        # Parser le fichier SIGOBE avec le service
        Result, Metadatafile, ColsToKeep = SigobeService.parse_fichier_excel(excel_file, annee, trimestre)

        logger.info(f"✅ Parsing réussi : {len(Result)} lignes à importer")

        # Déterminer le libellé de période
        if trimestre:
            periode_libelle = f"T{trimestre} {annee}"
        else:
            periode_libelle = f"Annuel {annee}"

        # Sauvegarder le fichier physiquement (SEULEMENT si parsing OK)
        from app.core.path_config import path_config

        relative_path = f"sigobe/{annee}/{fichier.filename}"
        upload_dir = path_config.UPLOADS_DIR / "sigobe" / str(annee)
        upload_dir.mkdir(parents=True, exist_ok=True)

        file_path = upload_dir / fichier.filename
        with open(file_path, "wb") as f:
            f.write(content)

        logger.info(f"📁 Fichier sauvegardé : {file_path}")

        # Générer l'URL avec ROOT_PATH
        file_url = path_config.get_file_url("uploads", relative_path)

        # 10. Créer l'enregistrement de chargement
        chargement = SigobeChargement(
            annee=annee,
            trimestre=trimestre,
            periode_libelle=periode_libelle,
            nom_fichier=fichier.filename,
            taille_octets=len(content),
            chemin_fichier=file_url,
            uploaded_by_user_id=current_user.id,
            statut="En cours",
        )

        session.add(chargement)
        session.commit()
        session.refresh(chargement)

        logger.info(f"✅ Chargement créé : ID={chargement.id}")

        # 11. Importer les lignes d'exécution depuis Result
        nb_lignes = 0
        programmes_set = set()
        actions_set = set()

        financial_columns = [
            "Budget_Vote",
            "Budget_Actuel",
            "Engagements_Emis",
            "Disponible_Eng",
            "Mandats_Emis",
            "Mandats_Vise_CF",
            "Mandats_Pec",
        ]

        for idx, row in Result.iterrows():
            try:
                # Extraire les montants depuis Result (colonnes typées)
                montants = {}
                for col in financial_columns:
                    if col in Result.columns:
                        val = row[col]
                        if pd.notna(val):
                            try:
                                montants[col.lower()] = Decimal(str(val))
                            except:
                                montants[col.lower()] = Decimal("0")
                        else:
                            montants[col.lower()] = Decimal("0")
                    else:
                        montants[col.lower()] = Decimal("0")

                # Extraire les métadonnées de la ligne (si présentes)
                periode_val = row.get("Periode") if "Periode" in Result.columns else None
                section_val = row.get("Section", "") if "Section" in Result.columns else Metadatafile.get("section", "")
                categorie_val = (
                    row.get("Categorie", "") if "Categorie" in Result.columns else Metadatafile.get("categorie", "")
                )
                type_credit_val = (
                    row.get("Type_credit", "")
                    if "Type_credit" in Result.columns
                    else Metadatafile.get("type_credit", "")
                )

                # Créer la ligne d'exécution
                execution = SigobeExecution(
                    chargement_id=chargement.id,
                    annee=annee,
                    trimestre=trimestre,
                    periode=periode_val if isinstance(periode_val, date) else None,
                    section=str(section_val) if section_val else None,
                    categorie=str(categorie_val) if categorie_val else None,
                    type_credit=str(type_credit_val) if type_credit_val else None,
                    programmes=str(row.get("Programmes", "")) if "Programmes" in Result.columns else "",
                    actions=str(row.get("Actions", "")) if "Actions" in Result.columns else "",
                    rprog=str(row.get("Rprog", "")) if "Rprog" in Result.columns else "",
                    type_depense=str(row.get("Type_depense", "")) if "Type_depense" in Result.columns else "",
                    activites=str(row.get("Activites", "")) if "Activites" in Result.columns else "",
                    taches=str(row.get("Taches", "")) if "Taches" in Result.columns else "",
                    budget_vote=montants.get("budget_vote", 0),
                    budget_actuel=montants.get("budget_actuel", 0),
                    engagements_emis=montants.get("engagements_emis", 0),
                    disponible_eng=montants.get("disponible_eng", 0),
                    mandats_emis=montants.get("mandats_emis", 0),
                    mandats_vise_cf=montants.get("mandats_vise_cf", 0),
                    mandats_pec=montants.get("mandats_pec", 0),
                )

                session.add(execution)
                nb_lignes += 1

                # Collecter les programmes/actions uniques
                if row.get("Programmes"):
                    programmes_set.add(str(row.get("Programmes")))
                if row.get("Actions"):
                    actions_set.add(str(row.get("Actions")))

                # Commit par batch de 100
                if nb_lignes % 100 == 0:
                    session.commit()
                    logger.info(f"💾 {nb_lignes} lignes importées...")

            except Exception as e:
                logger.warning(f"⚠️ Erreur ligne {idx}: {e}")
                continue

        # Commit final
        session.commit()

        # 12. Mettre à jour le chargement
        chargement.nb_lignes_importees = nb_lignes
        chargement.nb_programmes = len(programmes_set)
        chargement.nb_actions = len(actions_set)
        chargement.statut = "Terminé"
        session.add(chargement)
        session.commit()

        logger.info(
            f"✅ Import terminé : {nb_lignes} lignes, {len(programmes_set)} programmes, {len(actions_set)} actions"
        )

        # 13. Calculer les KPIs
        try:
            calcul_kpis_sigobe(chargement.id, session)
        except Exception as e:
            logger.error(f"❌ Erreur calcul KPIs : {e}")

        # Log activité
        ActivityService.log_user_activity(
            session=session,
            user=current_user,
            action_type="upload",
            target_type="sigobe",
            description=f"Import SIGOBE {periode_libelle} - {nb_lignes} lignes, {len(programmes_set)} programmes",
            target_id=chargement.id,
            icon="📊",
        )

        return {
            "ok": True,
            "chargement_id": chargement.id,
            "nb_lignes": nb_lignes,
            "nb_programmes": len(programmes_set),
            "nb_actions": len(actions_set),
            "message": f"Import réussi : {nb_lignes} lignes chargées",
        }

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"❌ Erreur upload SIGOBE : {e}")

        # Marquer le chargement comme en erreur si créé
        if "chargement" in locals():
            chargement.statut = "Erreur"
            chargement.message_erreur = str(e)
            session.add(chargement)
            session.commit()

        raise HTTPException(500, f"Erreur lors de l'import : {e!s}")


def calcul_kpis_sigobe(chargement_id: int, session: Session):
    """Calcule les KPIs agrégés pour un chargement SIGOBE"""

    chargement = session.get(SigobeChargement, chargement_id)
    if not chargement:
        return

    # Récupérer toutes les données d'exécution
    executions = session.exec(select(SigobeExecution).where(SigobeExecution.chargement_id == chargement_id)).all()

    if not executions:
        return

    # 1. KPI Global
    global_budget_vote = sum(e.budget_vote or 0 for e in executions)
    global_budget_actuel = sum(e.budget_actuel or 0 for e in executions)
    global_engagements = sum(e.engagements_emis or 0 for e in executions)
    global_mandats = sum(e.mandats_emis or 0 for e in executions)

    taux_engagement = (float(global_engagements) / float(global_budget_actuel) * 100) if global_budget_actuel > 0 else 0
    taux_mandatement = (float(global_mandats) / float(global_engagements) * 100) if global_engagements > 0 else 0
    taux_execution = (float(global_mandats) / float(global_budget_actuel) * 100) if global_budget_actuel > 0 else 0

    kpi_global = SigobeKpi(
        annee=chargement.annee,
        trimestre=chargement.trimestre,
        dimension="global",
        budget_vote_total=Decimal(str(global_budget_vote)),
        budget_actuel_total=Decimal(str(global_budget_actuel)),
        engagements_total=Decimal(str(global_engagements)),
        mandats_total=Decimal(str(global_mandats)),
        taux_engagement=Decimal(str(round(taux_engagement, 2))),
        taux_mandatement=Decimal(str(round(taux_mandatement, 2))),
        taux_execution=Decimal(str(round(taux_execution, 2))),
        chargement_id=chargement_id,
    )

    session.add(kpi_global)

    # 2. KPIs par programme
    programmes_dict = defaultdict(lambda: {"budget_vote": 0, "budget_actuel": 0, "engagements": 0, "mandats": 0})

    for e in executions:
        if e.programmes:
            programmes_dict[e.programmes]["budget_vote"] += e.budget_vote or 0
            programmes_dict[e.programmes]["budget_actuel"] += e.budget_actuel or 0
            programmes_dict[e.programmes]["engagements"] += e.engagements_emis or 0
            programmes_dict[e.programmes]["mandats"] += e.mandats_emis or 0

    for prog, data in programmes_dict.items():
        taux_eng = (float(data["engagements"]) / float(data["budget_actuel"]) * 100) if data["budget_actuel"] > 0 else 0
        taux_mand = (float(data["mandats"]) / float(data["engagements"]) * 100) if data["engagements"] > 0 else 0
        taux_exec = (float(data["mandats"]) / float(data["budget_actuel"]) * 100) if data["budget_actuel"] > 0 else 0

        code, libelle = split_code_libelle(prog)

        kpi_prog = SigobeKpi(
            annee=chargement.annee,
            trimestre=chargement.trimestre,
            dimension="programme",
            dimension_code=code,
            dimension_libelle=libelle,
            budget_vote_total=Decimal(str(data["budget_vote"])),
            budget_actuel_total=Decimal(str(data["budget_actuel"])),
            engagements_total=Decimal(str(data["engagements"])),
            mandats_total=Decimal(str(data["mandats"])),
            taux_engagement=Decimal(str(round(taux_eng, 2))),
            taux_mandatement=Decimal(str(round(taux_mand, 2))),
            taux_execution=Decimal(str(round(taux_exec, 2))),
            chargement_id=chargement_id,
        )

        session.add(kpi_prog)

    # 3. KPIs par nature de dépense
    natures_dict = defaultdict(lambda: {"budget_vote": 0, "budget_actuel": 0, "engagements": 0, "mandats": 0})

    for e in executions:
        if e.type_depense:
            natures_dict[e.type_depense]["budget_vote"] += e.budget_vote or 0
            natures_dict[e.type_depense]["budget_actuel"] += e.budget_actuel or 0
            natures_dict[e.type_depense]["engagements"] += e.engagements_emis or 0
            natures_dict[e.type_depense]["mandats"] += e.mandats_emis or 0

    for nature, data in natures_dict.items():
        taux_eng = (float(data["engagements"]) / float(data["budget_actuel"]) * 100) if data["budget_actuel"] > 0 else 0
        taux_mand = (float(data["mandats"]) / float(data["engagements"]) * 100) if data["engagements"] > 0 else 0
        taux_exec = (float(data["mandats"]) / float(data["budget_actuel"]) * 100) if data["budget_actuel"] > 0 else 0

        # Extraire le code court de la nature de dépense
        code_nature, libelle_nature = split_code_libelle(nature)

        # Si le code est vide ou identique au libellé, utiliser des abréviations standards
        if not code_nature or code_nature == libelle_nature:
            nature_lower = nature.lower()
            if "bien" in nature_lower or "service" in nature_lower:
                code_nature = "BS"
            elif "personnel" in nature_lower:
                code_nature = "P"
            elif "investissement" in nature_lower:
                code_nature = "I"
            elif "transfert" in nature_lower:
                code_nature = "T"
            else:
                code_nature = nature[:3].upper()  # Prendre les 3 premières lettres
            libelle_nature = nature

        kpi_nature = SigobeKpi(
            annee=chargement.annee,
            trimestre=chargement.trimestre,
            dimension="nature",
            dimension_code=code_nature,
            dimension_libelle=libelle_nature,
            budget_vote_total=Decimal(str(data["budget_vote"])),
            budget_actuel_total=Decimal(str(data["budget_actuel"])),
            engagements_total=Decimal(str(data["engagements"])),
            mandats_total=Decimal(str(data["mandats"])),
            taux_engagement=Decimal(str(round(taux_eng, 2))),
            taux_mandatement=Decimal(str(round(taux_mand, 2))),
            taux_execution=Decimal(str(round(taux_exec, 2))),
            chargement_id=chargement_id,
        )

        session.add(kpi_nature)

    session.commit()
    logger.info(f"✅ KPIs calculés : 1 global + {len(programmes_dict)} programmes + {len(natures_dict)} natures")


@router.delete("/api/sigobe/{chargement_id}")
def api_delete_sigobe(
    chargement_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """Supprimer un chargement SIGOBE et toutes ses données"""
    chargement = session.get(SigobeChargement, chargement_id)
    if not chargement:
        raise HTTPException(404, "Chargement non trouvé")

    try:
        # Supprimer les KPIs
        session.exec(delete(SigobeKpi).where(SigobeKpi.chargement_id == chargement_id))

        # Supprimer les exécutions
        session.exec(delete(SigobeExecution).where(SigobeExecution.chargement_id == chargement_id))

        # Supprimer le chargement
        periode_libelle = chargement.periode_libelle
        session.delete(chargement)
        session.commit()

        logger.info(f"✅ Chargement SIGOBE {chargement_id} supprimé par {current_user.email}")

        # Log activité
        ActivityService.log_user_activity(
            session=session,
            user=current_user,
            action_type="delete",
            target_type="sigobe",
            description=f"Suppression du chargement SIGOBE {periode_libelle}",
            target_id=chargement_id,
            icon="🗑️",
        )

        return {"ok": True, "message": "Chargement supprimé avec succès"}

    except Exception as e:
        logger.error(f"❌ Erreur suppression SIGOBE {chargement_id}: {e}")
        session.rollback()
        raise HTTPException(500, f"Erreur lors de la suppression : {e!s}")


@router.get("/api/sigobe/executions/{execution_id}")
def api_get_sigobe_execution(
    execution_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """Récupérer une ligne d'exécution SIGOBE"""
    execution = session.get(SigobeExecution, execution_id)
    if not execution:
        raise HTTPException(404, "Ligne d'exécution non trouvée")

    return {
        "ok": True,
        "execution": {
            "id": execution.id,
            "programmes": execution.programmes,
            "actions": execution.actions,
            "rprog": execution.rprog,
            "type_depense": execution.type_depense,
            "activites": execution.activites,
            "taches": execution.taches,
            "budget_vote": float(execution.budget_vote),
            "budget_actuel": float(execution.budget_actuel),
            "engagements_emis": float(execution.engagements_emis),
            "disponible_eng": float(execution.disponible_eng),
            "mandats_emis": float(execution.mandats_emis),
            "mandats_vise_cf": float(execution.mandats_vise_cf),
            "mandats_pec": float(execution.mandats_pec),
        },
    }


@router.put("/api/sigobe/executions/{execution_id}")
async def api_update_sigobe_execution(
    execution_id: int,
    data: dict,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Modifier une ligne d'exécution SIGOBE"""
    execution = session.get(SigobeExecution, execution_id)
    if not execution:
        raise HTTPException(404, "Ligne d'exécution non trouvée")

    try:
        # Mettre à jour les montants financiers
        execution.budget_vote = Decimal(str(data.get("budget_vote", 0)))
        execution.budget_actuel = Decimal(str(data.get("budget_actuel", 0)))
        execution.engagements_emis = Decimal(str(data.get("engagements_emis", 0)))
        execution.disponible_eng = Decimal(str(data.get("disponible_eng", 0)))
        execution.mandats_emis = Decimal(str(data.get("mandats_emis", 0)))
        execution.mandats_vise_cf = Decimal(str(data.get("mandats_vise_cf", 0)))
        execution.mandats_pec = Decimal(str(data.get("mandats_pec", 0)))

        session.add(execution)
        session.commit()

        logger.info(f"✅ Ligne SIGOBE {execution_id} modifiée par {current_user.email}")

        # Log activité
        ActivityService.log_user_activity(
            session=session,
            user=current_user,
            action_type="update",
            target_type="sigobe_execution",
            description=f"Modification ligne SIGOBE {execution.taches or execution.activites or execution.actions}",
            target_id=execution_id,
            icon="✏️",
        )

        return {"ok": True, "message": "Ligne modifiée avec succès"}

    except Exception as e:
        logger.error(f"❌ Erreur modification ligne SIGOBE {execution_id}: {e}")
        session.rollback()
        raise HTTPException(500, f"Erreur lors de la modification : {e!s}")


@router.delete("/api/sigobe/executions/{execution_id}")
def api_delete_sigobe_execution(
    execution_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """Supprimer une ligne d'exécution SIGOBE"""
    execution = session.get(SigobeExecution, execution_id)
    if not execution:
        raise HTTPException(404, "Ligne d'exécution non trouvée")

    try:
        tache_libelle = execution.taches or execution.activites or execution.actions or "ligne"

        session.delete(execution)
        session.commit()

        logger.info(f"✅ Ligne SIGOBE {execution_id} supprimée par {current_user.email}")

        # Log activité
        ActivityService.log_user_activity(
            session=session,
            user=current_user,
            action_type="delete",
            target_type="sigobe_execution",
            description=f"Suppression ligne SIGOBE {tache_libelle}",
            target_id=execution_id,
            icon="🗑️",
        )

        return {"ok": True, "message": "Ligne supprimée avec succès"}

    except Exception as e:
        session.rollback()
        logger.error(f"❌ Erreur suppression ligne SIGOBE {execution_id}: {e}")
        raise HTTPException(500, f"Erreur lors de la suppression: {e!s}")


@router.get("/api/sigobe/{chargement_id}/export/excel")
def api_export_sigobe_excel(
    chargement_id: int, session: Session = Depends(get_session), current_user: User = Depends(get_current_user)
):
    """
    Exporter un chargement SIGOBE en Excel avec la même mise en forme que le template
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    except ImportError:
        raise HTTPException(500, "openpyxl n'est pas installé")

    # Récupérer le chargement
    chargement = session.get(SigobeChargement, chargement_id)
    if not chargement:
        raise HTTPException(404, "Chargement SIGOBE non trouvé")

    # Récupérer toutes les lignes d'exécution
    executions = session.exec(
        select(SigobeExecution)
        .where(SigobeExecution.chargement_id == chargement_id)
        .order_by(
            SigobeExecution.programmes, SigobeExecution.actions, SigobeExecution.activites, SigobeExecution.taches
        )
    ).all()

    # Créer le classeur Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"SIGOBE {chargement.periode_libelle}"

    # Styles
    header_fill = PatternFill(start_color="FF8C00", end_color="FF8C00", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    data_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    total_action_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    total_programme_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
    total_general_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    total_font = Font(bold=True, size=10)
    total_general_font = Font(bold=True, size=11, color="FFFFFF")
    border_style = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    # En-têtes
    headers = [
        "PROGRAMMES",
        "ACTIONS",
        "RPROG",
        "TYPE DEPENSE",
        "ACTIVITES",
        "TACHES",
        "BUDGET VOTE",
        "BUDGET ACTUEL",
        "ENGAGEMENTS EMIS",
        "DISPONIBLE ENG",
        "MANDATS EMIS",
        "MANDATS VISE CF",
        "MANDATS PEC",
    ]

    current_row = 1
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_style

    # Ajuster largeurs
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col].width = 30
    for col in ["G", "H", "I", "J", "K", "L", "M"]:
        ws.column_dimensions[col].width = 18

    current_row = 2

    # Organiser par hiérarchie
    from collections import defaultdict

    programmes_dict = defaultdict(lambda: defaultdict(list))

    for exec in executions:
        prog = exec.programmes or "Sans programme"
        action = exec.actions or "Sans action"
        programmes_dict[prog][action].append(exec)

    # Totaux globaux
    totaux_generaux = {
        "vote": 0,
        "actuel": 0,
        "engagements": 0,
        "disponible": 0,
        "mandats": 0,
        "mandats_vise": 0,
        "mandats_pec": 0,
    }

    # Écrire les données avec hiérarchie
    for programme, actions_dict in sorted(programmes_dict.items()):
        # Ligne Programme
        cell = ws.cell(row=current_row, column=1)
        cell.value = programme
        cell.fill = total_programme_fill
        cell.font = total_font
        cell.border = border_style
        for col in range(2, 14):
            ws.cell(row=current_row, column=col).border = border_style
            ws.cell(row=current_row, column=col).fill = total_programme_fill
        current_row += 1

        totaux_programme = {
            "vote": 0,
            "actuel": 0,
            "engagements": 0,
            "disponible": 0,
            "mandats": 0,
            "mandats_vise": 0,
            "mandats_pec": 0,
        }

        for action, execs in sorted(actions_dict.items()):
            # Ligne Action
            cell = ws.cell(row=current_row, column=2)
            cell.value = action
            cell.fill = total_action_fill
            cell.font = total_font
            cell.border = border_style
            ws.cell(row=current_row, column=1).border = border_style
            for col in range(3, 14):
                ws.cell(row=current_row, column=col).border = border_style
                ws.cell(row=current_row, column=col).fill = total_action_fill
            current_row += 1

            # Lignes d'exécution (tâches)
            for exec in execs:
                row_data = [
                    exec.programmes or "",
                    exec.actions or "",
                    exec.rprog or "",
                    exec.type_depense or "",
                    exec.activites or "",
                    exec.taches or "",
                    float(exec.budget_vote or 0),
                    float(exec.budget_actuel or 0),
                    float(exec.engagements_emis or 0),
                    float(exec.disponible_eng or 0),
                    float(exec.mandats_emis or 0),
                    float(exec.mandats_vise_cf or 0),
                    float(exec.mandats_pec or 0),
                ]

                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col_num)
                    cell.value = value
                    cell.border = border_style

                    if col_num > 6:  # Colonnes financières
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                        cell.number_format = "#,##0"
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

                # Accumuler totaux
                totaux_programme["vote"] += float(exec.budget_vote or 0)
                totaux_programme["actuel"] += float(exec.budget_actuel or 0)
                totaux_programme["engagements"] += float(exec.engagements_emis or 0)
                totaux_programme["disponible"] += float(exec.disponible_eng or 0)
                totaux_programme["mandats"] += float(exec.mandats_emis or 0)
                totaux_programme["mandats_vise"] += float(exec.mandats_vise_cf or 0)
                totaux_programme["mandats_pec"] += float(exec.mandats_pec or 0)

                current_row += 1

        # Total Programme
        row_data = [
            f"TOTAL {programme}",
            "",
            "",
            "",
            "",
            "",
            totaux_programme["vote"],
            totaux_programme["actuel"],
            totaux_programme["engagements"],
            totaux_programme["disponible"],
            totaux_programme["mandats"],
            totaux_programme["mandats_vise"],
            totaux_programme["mandats_pec"],
        ]

        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = value
            cell.fill = total_programme_fill
            cell.font = total_font
            cell.border = border_style
            if col_num > 6:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        # Accumuler totaux généraux
        for key in totaux_generaux:
            totaux_generaux[key] += totaux_programme[key]

        current_row += 1
        current_row += 1  # Ligne vide

    # Total Général
    row_data = [
        "TOTAL GÉNÉRAL",
        "",
        "",
        "",
        "",
        "",
        totaux_generaux["vote"],
        totaux_generaux["actuel"],
        totaux_generaux["engagements"],
        totaux_generaux["disponible"],
        totaux_generaux["mandats"],
        totaux_generaux["mandats_vise"],
        totaux_generaux["mandats_pec"],
    ]

    for col_num, value in enumerate(row_data, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = value
        cell.fill = total_general_fill
        cell.font = total_general_font
        cell.border = border_style
        if col_num > 6:
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.number_format = "#,##0"
        else:
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Figer les en-têtes
    ws.freeze_panes = "A2"

    # Sauvegarder dans un buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(f"✅ Excel exporté pour chargement SIGOBE {chargement_id}")

    # Nom du fichier
    filename = f"SIGOBE_{chargement.periode_libelle.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/api/sigobe/{chargement_id}/kpis")
def api_get_kpis_sigobe(
    chargement_id: int,
    dimension: str | None = None,
    session: Session = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Récupérer les KPIs d'un chargement SIGOBE"""

    query = select(SigobeKpi).where(SigobeKpi.chargement_id == chargement_id)

    if dimension:
        query = query.where(SigobeKpi.dimension == dimension)

    kpis = session.exec(query).all()

    return {
        "ok": True,
        "kpis": [
            {
                "id": kpi.id,
                "dimension": kpi.dimension,
                "dimension_code": kpi.dimension_code,
                "dimension_libelle": kpi.dimension_libelle,
                "budget_vote_total": float(kpi.budget_vote_total),
                "budget_actuel_total": float(kpi.budget_actuel_total),
                "engagements_total": float(kpi.engagements_total),
                "mandats_total": float(kpi.mandats_total),
                "taux_engagement": float(kpi.taux_engagement),
                "taux_mandatement": float(kpi.taux_mandatement),
                "taux_execution": float(kpi.taux_execution),
            }
            for kpi in kpis
        ],
    }
